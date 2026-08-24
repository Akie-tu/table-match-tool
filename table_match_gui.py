#!/usr/bin/env python3
"""
电商工具 v6.0 - 表格核对 + 图片转换 + 开票生成
============================================
Tab1 表格核对: 用源表某字段去目标表检索, 回填字段 + 多规格标记
Tab2 图片转换: 批量图片转JPG(webp/png/heic等→jpg), 保留目录树
Tab3 开票生成: 数电发票批量开票模板生成(流水号自动+固定内容+手动/批量录入)

功能:
  - 任意检索键列(快递单号/订单号等)
  - 增量填充(已有数据不覆盖)
  - 多规格标记(≥2种商品→备注"多规格")
  - 输出另存加后缀(不修改源文件)
  - 图片批量转JPG, 保留目录结构
  - GUI(tkinter) + CLI 双界面
纯 tkinter 标准库 + openpyxl + Pillow
"""
import os
import sys
import shutil
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from collections import defaultdict

import openpyxl

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

try:
    from invoice_gen import (generate_invoice_xlsx, parse_bulk_text,
                              DEFAULT_FIXED, INVOICE_TYPE_OPTIONS, find_template)
    INVOICE_AVAILABLE = True
except ImportError:
    INVOICE_AVAILABLE = False


# ================= 表格核对核心 =================
def find_header_row(ws, keywords):
    for r in range(1, min(6, ws.max_row + 1)):
        strs = [str(ws.cell(row=r, column=c + 1).value or '').strip()
                for c in range(ws.max_column)]
        if sum(1 for k in keywords if any(k.lower() in s.lower() for s in strs)) >= len(keywords):
            return r
    return None


def col_index_by_header(ws, header_row, header_name):
    for c in range(ws.max_column):
        if str(ws.cell(row=header_row, column=c + 1).value or '').strip() == header_name:
            return c
    return None


def run_match(src_path, tgt_path, src_key, tgt_key, fill_map, sku_col, remark_col,
              skip_existing, progress_cb=None):
    twb = openpyxl.load_workbook(tgt_path, data_only=True)
    tws = twb[twb.sheetnames[0]]
    thr = find_header_row(tws, [tgt_key])
    if thr is None:
        raise ValueError(f"目标表找不到表头(含 '{tgt_key}')")
    key_col = col_index_by_header(tws, thr, tgt_key)
    if key_col is None:
        raise ValueError(f"目标表没有列 '{tgt_key}'")
    tgt_cols = {}
    for sc, tc in fill_map:
        ci = col_index_by_header(tws, thr, tc)
        if ci is not None:
            tgt_cols.setdefault(sc, []).append(ci)
    sku_idx = col_index_by_header(tws, thr, sku_col) if sku_col else None

    idx = defaultdict(list)
    for r in tws.iter_rows(min_row=thr + 1, values_only=True):
        k = r[key_col]
        if not k:
            continue
        for part in str(k).replace('，', ',').replace(' ', '').split(','):
            if part.strip():
                idx[part.strip()].append(r)

    swb = openpyxl.load_workbook(src_path)
    sws = swb[swb.sheetnames[0]]
    shr = find_header_row(sws, [src_key])
    if shr is None:
        raise ValueError(f"源表找不到表头(含 '{src_key}')")
    src_key_col = col_index_by_header(sws, shr, src_key)
    src_fill_cols = {}
    for sc, tc in fill_map:
        ci = col_index_by_header(sws, shr, sc)
        if ci is not None:
            src_fill_cols[sc] = ci
    remark_col_idx = col_index_by_header(sws, shr, remark_col) if remark_col else None

    matched = 0
    multi = 0
    notfound = []
    total = 0
    for row in sws.iter_rows(min_row=shr + 1):
        k = row[src_key_col].value
        if not k:
            continue
        total += 1
        k = str(k).strip()
        hits = idx.get(k)
        if hits:
            matched += 1
            for sc, tc in fill_map:
                sci = src_fill_cols.get(sc)
                tci_list = tgt_cols.get(sc)
                if sci is not None and tci_list and not row[sci].value:
                    row[sci].value = hits[0][tci_list[0]]
            if remark_col_idx is not None and not row[remark_col_idx].value:
                if sku_idx is not None and len(hits) >= 2:
                    uniq = {str(h[sku_idx] or '').strip() for h in hits if h[sku_idx]}
                    if len(uniq) >= 2:
                        row[remark_col_idx].value = "多规格"
                        multi += 1
        else:
            notfound.append(k)
        if progress_cb and total % 10 == 0:
            progress_cb(total)

    base, ext = os.path.splitext(src_path)
    out_path = f"{base}_结果{ext}"
    swb.save(out_path)
    return matched, multi, notfound, out_path


# ================= 图片转换核心 =================
IMG_EXTS = {'.webp', '.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff', '.heic', '.avif', '.jfif'}
QUALITY = 92


def convert_one(src, dst):
    img = Image.open(src)
    if img.mode in ('RGBA', 'LA', 'P', 'PA'):
        rgba = img.convert('RGBA')
        bg = Image.new('RGB', rgba.size, (255, 255, 255))
        bg.paste(rgba, mask=rgba.split()[-1])
        img = bg
    elif img.mode != 'RGB':
        img = img.convert('RGB')
    img.save(dst, 'JPEG', quality=QUALITY)


def run_imgconvert(src_root, out_root, progress_cb=None):
    """批量转JPG, 保留目录树. 返回 (total, converted, copied, failed, errors)"""
    if not PIL_AVAILABLE:
        raise ValueError("缺少 Pillow 库, 无法转换图片. 请运行: pip install pillow")
    if not os.path.isdir(src_root):
        raise ValueError(f"路径不存在: {src_root}")
    os.makedirs(out_root, exist_ok=True)
    total = converted = copied = failed = 0
    errors = []
    abs_out = os.path.abspath(out_root)
    count = 0
    for root, dirs, files in os.walk(src_root):
        if os.path.abspath(root).startswith(abs_out):
            continue
        for f in files:
            src = os.path.join(root, f)
            rel = os.path.relpath(src, src_root)
            total += 1
            try:
                if os.path.splitext(f)[1].lower() in IMG_EXTS:
                    base = os.path.splitext(f)[0]
                    dst = os.path.join(out_root, os.path.dirname(rel), base + ".jpg")
                    os.makedirs(os.path.dirname(dst), exist_ok=True)
                    convert_one(src, dst)
                    converted += 1
                else:
                    dst = os.path.join(out_root, rel)
                    os.makedirs(os.path.dirname(dst), exist_ok=True)
                    shutil.copy2(src, dst)
                    copied += 1
            except Exception as e:
                failed += 1
                errors.append((rel, str(e)))
            count += 1
            if progress_cb and count % 20 == 0:
                progress_cb(count, converted, failed)
    return total, converted, copied, failed, errors


# ================= GUI =================
class App:
    def __init__(self, root):
        self.root = root
        root.title("电商工具 v6.0")
        root.geometry("780x700")
        root.minsize(700, 620)

        self.nb = ttk.Notebook(root)
        self.nb.pack(fill="both", expand=True, padx=6, pady=6)

        self.build_match_tab()
        self.build_img_tab()
        self.build_invoice_tab()

    # ---------- Tab1 表格核对 ----------
    def build_match_tab(self):
        tab = ttk.Frame(self.nb)
        self.nb.add(tab, text="① 表格核对")

        frm = ttk.LabelFrame(tab, text="1. 选择文件")
        frm.pack(fill="x", padx=10, pady=5)
        self.src_var = tk.StringVar()
        self.tgt_var = tk.StringVar()
        ttk.Label(frm, text="源表(被填的):").grid(row=0, column=0, sticky="w", padx=5, pady=3)
        ttk.Entry(frm, textvariable=self.src_var, width=45).grid(row=0, column=1, padx=5)
        ttk.Button(frm, text="浏览", command=lambda: self.pick(self.src_var)).grid(row=0, column=2)
        ttk.Label(frm, text="目标表(数据源):").grid(row=1, column=0, sticky="w", padx=5)
        ttk.Entry(frm, textvariable=self.tgt_var, width=45).grid(row=1, column=1, padx=5)
        ttk.Button(frm, text="浏览", command=lambda: self.pick(self.tgt_var)).grid(row=1, column=2)

        frm2 = ttk.LabelFrame(tab, text="2. 检索键列 (表头名)")
        frm2.pack(fill="x", padx=10, pady=5)
        self.sk = tk.StringVar(value="快递单号")
        self.tk = tk.StringVar(value="快递单号")
        ttk.Label(frm2, text="源表键列:").grid(row=0, column=0, padx=5)
        ttk.Entry(frm2, textvariable=self.sk, width=20).grid(row=0, column=1, padx=5)
        ttk.Label(frm2, text="目标表键列:").grid(row=0, column=2, padx=5)
        ttk.Entry(frm2, textvariable=self.tk, width=20).grid(row=0, column=3, padx=5)

        frm3 = ttk.LabelFrame(tab, text="3. 回填映射 (源列 → 目标列)")
        frm3.pack(fill="x", padx=10, pady=5)
        self.map_rows = []
        self._add_row(frm3)
        self.add_btn = ttk.Button(frm3, text="+ 添加映射", command=lambda: self._add_row(frm3))
        self.add_btn.pack(anchor="w", padx=10, pady=3)

        frm4 = ttk.LabelFrame(tab, text="4. 多规格标记 (可空)")
        frm4.pack(fill="x", padx=10, pady=5)
        self.sku_col = tk.StringVar()
        self.rmk_col = tk.StringVar()
        ttk.Label(frm4, text="多规格判断列:").grid(row=0, column=0, padx=5)
        ttk.Entry(frm4, textvariable=self.sku_col, width=18).grid(row=0, column=1, padx=5)
        ttk.Label(frm4, text="备注列:").grid(row=0, column=2, padx=5)
        ttk.Entry(frm4, textvariable=self.rmk_col, width=18).grid(row=0, column=3, padx=5)

        frm5 = ttk.Frame(tab)
        frm5.pack(fill="x", padx=10, pady=5)
        self.skip_existing = tk.BooleanVar(value=True)
        ttk.Checkbutton(frm5, text="已有数据不覆盖", variable=self.skip_existing).pack(side="left", padx=5)
        self.run_btn = ttk.Button(frm5, text="▶ 开始核对", command=self.run)
        self.run_btn.pack(side="right", padx=5)

        self.match_log = scrolledtext.ScrolledText(tab, height=12, font=("Consolas", 9))
        self.match_log.pack(fill="both", expand=True, padx=10, pady=5)

    # ---------- Tab2 图片转换 ----------
    def build_img_tab(self):
        tab = ttk.Frame(self.nb)
        self.nb.add(tab, text="② 图片转JPG")

        frm = ttk.LabelFrame(tab, text="1. 选择图片文件夹")
        frm.pack(fill="x", padx=10, pady=5)
        self.img_src = tk.StringVar()
        self.img_out = tk.StringVar()
        ttk.Label(frm, text="源文件夹:").grid(row=0, column=0, sticky="w", padx=5, pady=3)
        ttk.Entry(frm, textvariable=self.img_src, width=45).grid(row=0, column=1, padx=5)
        ttk.Button(frm, text="浏览", command=lambda: self.pick_dir(self.img_src)).grid(row=0, column=2)
        ttk.Label(frm, text="输出路径(可空=自动):").grid(row=1, column=0, sticky="w", padx=5)
        ttk.Entry(frm, textvariable=self.img_out, width=45).grid(row=1, column=1, padx=5)
        ttk.Button(frm, text="浏览", command=lambda: self.pick_dir(self.img_out)).grid(row=1, column=2)

        frm2 = ttk.Frame(tab)
        frm2.pack(fill="x", padx=10, pady=5)
        ttk.Label(frm2, text="质量 (1-100):").pack(side="left")
        self.img_q = tk.IntVar(value=92)
        ttk.Spinbox(frm2, from_=1, to=100, textvariable=self.img_q, width=5).pack(side="left", padx=5)
        self.img_btn = ttk.Button(frm2, text="▶ 开始转换", command=self.run_img)
        self.img_btn.pack(side="right", padx=5)

        self.img_log = scrolledtext.ScrolledText(tab, height=12, font=("Consolas", 9))
        self.img_log.pack(fill="both", expand=True, padx=10, pady=5)

    # ---------- Tab3 开票生成 ----------
    def build_invoice_tab(self):
        tab = ttk.Frame(self.nb)
        self.nb.add(tab, text="③ 开票生成")

        # 0. 固定内容配置
        frm0 = ttk.LabelFrame(tab, text="固定内容 (可修改, 通常不用动)")
        frm0.pack(fill="x", padx=10, pady=5)
        self.inv_type = tk.StringVar(value=DEFAULT_FIXED["invoice_type"])
        self.inv_taxinc = tk.StringVar(value=DEFAULT_FIXED["tax_included"])
        self.inv_item = tk.StringVar(value=DEFAULT_FIXED["item_name"])
        self.inv_code = tk.StringVar(value=DEFAULT_FIXED["tax_code"])
        self.inv_unit = tk.StringVar(value=DEFAULT_FIXED["unit"])
        self.inv_rate = tk.StringVar(value=DEFAULT_FIXED["tax_rate"])
        ttk.Label(frm0, text="发票类型:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        ttk.Combobox(frm0, textvariable=self.inv_type, values=INVOICE_TYPE_OPTIONS,
                     width=12, state="readonly").grid(row=0, column=1, padx=5)
        ttk.Label(frm0, text="是否含税:").grid(row=0, column=2, sticky="w", padx=5)
        ttk.Combobox(frm0, textvariable=self.inv_taxinc, values=["是", "否"],
                     width=4, state="readonly").grid(row=0, column=3, padx=5)
        ttk.Label(frm0, text="项目名称:").grid(row=0, column=4, sticky="w", padx=5)
        ttk.Entry(frm0, textvariable=self.inv_item, width=10).grid(row=0, column=5, padx=5)
        ttk.Label(frm0, text="税收编码:").grid(row=0, column=6, sticky="w", padx=5)
        ttk.Entry(frm0, textvariable=self.inv_code, width=22).grid(row=0, column=7, padx=5)
        ttk.Label(frm0, text="单位:").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        ttk.Entry(frm0, textvariable=self.inv_unit, width=6).grid(row=1, column=1, padx=5, sticky="w")
        ttk.Label(frm0, text="税率:").grid(row=1, column=2, sticky="w", padx=5)
        ttk.Entry(frm0, textvariable=self.inv_rate, width=6).grid(row=1, column=3, padx=5, sticky="w")
        self.inv_tpl = tk.StringVar(value="")
        ttk.Label(frm0, text="模板:").grid(row=1, column=4, sticky="w", padx=5)
        ttk.Entry(frm0, textvariable=self.inv_tpl, width=26).grid(row=1, column=5, columnspan=2, padx=5)
        ttk.Button(frm0, text="浏览", command=lambda: self.pick(self.inv_tpl)).grid(row=1, column=7, padx=5)

        # 1. 手动录入
        frm1 = ttk.LabelFrame(tab, text="手动添加一行 (自然人或海外无税号 → 勾选'自然人')")
        frm1.pack(fill="x", padx=10, pady=5)
        self.iv_buyer = tk.StringVar()
        self.iv_taxid = tk.StringVar()
        self.iv_natural = tk.BooleanVar(value=False)
        self.iv_qty = tk.StringVar(value="1")
        self.iv_amt = tk.StringVar()
        self.iv_rmk = tk.StringVar()
        ttk.Label(frm1, text="购买方名称*:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        ttk.Entry(frm1, textvariable=self.iv_buyer, width=30).grid(row=0, column=1, padx=5)
        ttk.Label(frm1, text="税号(公司必填):").grid(row=0, column=2, sticky="w", padx=5)
        ttk.Entry(frm1, textvariable=self.iv_taxid, width=22).grid(row=0, column=3, padx=5)
        ttk.Checkbutton(frm1, text="自然人", variable=self.iv_natural).grid(row=0, column=4, padx=5)
        ttk.Label(frm1, text="数量:").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        ttk.Entry(frm1, textvariable=self.iv_qty, width=6).grid(row=1, column=1, padx=5, sticky="w")
        ttk.Label(frm1, text="金额*:").grid(row=1, column=2, sticky="w", padx=5)
        ttk.Entry(frm1, textvariable=self.iv_amt, width=12).grid(row=1, column=3, padx=5, sticky="w")
        ttk.Label(frm1, text="备注:").grid(row=1, column=4, sticky="w", padx=5)
        ttk.Entry(frm1, textvariable=self.iv_rmk, width=24).grid(row=1, column=5, padx=5)
        ttk.Button(frm1, text="＋ 添加", command=self.invoice_add_one).grid(row=1, column=6, padx=8)
        ttk.Button(frm1, text="批量导入…", command=self.invoice_bulk_dlg).grid(row=1, column=7, padx=5)

        # 2. 发票列表
        frm2 = ttk.LabelFrame(tab, text="发票列表 (流水号自动生成)")
        frm2.pack(fill="both", expand=True, padx=10, pady=5)
        cols = ("serial", "buyer", "taxid", "natural", "qty", "amount", "remark")
        self.inv_tree = ttk.Treeview(frm2, columns=cols, show="headings", height=8)
        headers = {"serial": "流水号", "buyer": "购买方名称", "taxid": "税号",
                   "natural": "自然人", "qty": "数量", "amount": "金额", "remark": "备注"}
        widths = {"serial": 55, "buyer": 200, "taxid": 170, "natural": 50,
                  "qty": 50, "amount": 80, "remark": 120}
        for c in cols:
            self.inv_tree.heading(c, text=headers[c])
            self.inv_tree.column(c, width=widths[c], anchor="w")
        self.inv_tree.pack(fill="both", expand=True, padx=5, pady=5)
        self.invoices = []  # 内存列表 [{buyer,tax_id,is_natural,qty,amount,remark}]

        # 3. 操作按钮
        frm3 = ttk.Frame(tab)
        frm3.pack(fill="x", padx=10, pady=5)
        ttk.Button(frm3, text="删除选中", command=self.invoice_del_sel).pack(side="left", padx=5)
        ttk.Button(frm3, text="清空列表", command=self.invoice_clear).pack(side="left", padx=5)
        ttk.Button(frm3, text="▶ 生成开票xlsx", command=self.invoice_generate).pack(side="right", padx=5)

        self.inv_log = scrolledtext.ScrolledText(tab, height=6, font=("Consolas", 9))
        self.inv_log.pack(fill="both", expand=True, padx=10, pady=5)

        if not INVOICE_AVAILABLE:
            self.log(self.inv_log, "⚠️ 开票模块未加载(invoice_gen.py缺失), 请联系管理员")

    def invoice_refresh_tree(self):
        self.inv_tree.delete(*self.inv_tree.get_children())
        for i, inv in enumerate(self.invoices, 1):
            sn = f"{i:03d}"
            self.inv_tree.insert("", "end", values=(
                sn, inv.get("buyer", ""), inv.get("tax_id", ""),
                inv.get("is_natural", "") or "", inv.get("qty", ""),
                inv.get("amount", ""), inv.get("remark", "")))

    def invoice_add_one(self):
        buyer = self.iv_buyer.get().strip()
        amt = self.iv_amt.get().strip()
        if not buyer:
            messagebox.showwarning("提示", "请填写购买方名称")
            return
        if not amt:
            messagebox.showwarning("提示", "请填写金额")
            return
        tax_id = self.iv_taxid.get().strip()
        inv = {
            "buyer": buyer, "tax_id": tax_id,
            "is_natural": "是" if self.iv_natural.get() else "",
            "qty": self.iv_qty.get().strip(), "amount": amt,
            "remark": self.iv_rmk.get().strip(),
        }
        self.invoices.append(inv)
        self.invoice_refresh_tree()
        # 清空输入(保留数量和自然人勾选, 名称/税号/金额/备注清掉)
        self.iv_buyer.set("")
        self.iv_taxid.set("")
        self.iv_amt.set("")
        self.iv_rmk.set("")
        self.log(self.inv_log, f"✔ 已添加: {buyer} 金额={amt}")

    def invoice_bulk_dlg(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("批量导入开票数据")
        dlg.geometry("560x420")
        dlg.transient(self.root)
        dlg.grab_set()
        ttk.Label(dlg, text="每行一条, 格式: 购买方名称, 税号或\"是\", 数量, 金额 (,备注可选)\n"
                            "例: 张三,是,2,52.20    公司A,91370306MA3TBJ0T8E,1,122.00,加急\n"
                            "海外/自然人无税号: 税号位写\"是\"或留空").pack(anchor="w", padx=10, pady=5)
        txt = scrolledtext.ScrolledText(dlg, height=14, font=("Consolas", 10))
        txt.pack(fill="both", expand=True, padx=10, pady=5)

        def do_import():
            try:
                lst = parse_bulk_text(txt.get("1.0", "end"))
            except ValueError as e:
                messagebox.showerror("格式错误", str(e), parent=dlg)
                return
            if not lst:
                messagebox.showwarning("提示", "没有解析到任何行", parent=dlg)
                return
            self.invoices.extend(lst)
            self.invoice_refresh_tree()
            self.log(self.inv_log, f"✔ 批量导入 {len(lst)} 条")
            dlg.destroy()

        frm = ttk.Frame(dlg)
        frm.pack(fill="x", padx=10, pady=8)
        ttk.Button(frm, text="导入", command=do_import).pack(side="right", padx=5)
        ttk.Button(frm, text="取消", command=dlg.destroy).pack(side="right", padx=5)

    def invoice_del_sel(self):
        sel = self.inv_tree.selection()
        if not sel:
            return
        idxs = sorted(int(self.inv_tree.item(i, "values")[0]) - 1 for i in sel)
        for idx in reversed(idxs):
            if 0 <= idx < len(self.invoices):
                del self.invoices[idx]
        self.invoice_refresh_tree()

    def invoice_clear(self):
        if self.invoices and messagebox.askyesno("确认", "清空全部发票?"):
            self.invoices.clear()
            self.invoice_refresh_tree()
            self.log(self.inv_log, "🗑 列表已清空")

    def invoice_generate(self):
        if not self.invoices:
            messagebox.showwarning("提示", "发票列表为空, 请先添加")
            return
        # 用当前固定内容覆盖默认
        fixed = {
            "invoice_type": self.inv_type.get(),
            "tax_included": self.inv_taxinc.get(),
            "item_name": self.inv_item.get().strip(),
            "tax_code": self.inv_code.get().strip(),
            "unit": self.inv_unit.get().strip(),
            "tax_rate": self.inv_rate.get().strip(),
        }
        invoices = [{**inv, **fixed} for inv in self.invoices]
        tpl = self.inv_tpl.get().strip().strip('"') or find_template()

        out = filedialog.asksaveasfilename(
            defaultextension=".xlsx", initialfile="开票导入.xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if not out:
            return
        self.inv_log.insert("end", "▶ 生成中...\n")
        try:
            path, errs = generate_invoice_xlsx(invoices, template_path=tpl, out_path=out)
            if errs:
                self.log(self.inv_log, f"❌ 校验未通过 ({len(errs)} 条):")
                for e in errs:
                    self.log(self.inv_log, f"   {e}")
                messagebox.showerror("校验失败", "请修正数据:\n" + "\n".join(errs[:10]))
                return
            self.log(self.inv_log, f"✅ 生成成功: {path}")
            self.log(self.inv_log, f"   共 {len(invoices)} 张发票, 流水号 001~{len(invoices):03d}")
            messagebox.showinfo("成功", f"开票文件已生成!\n{path}\n共 {len(invoices)} 张")
        except Exception as e:
            self.log(self.inv_log, f"❌ 错误: {e}")
            messagebox.showerror("错误", str(e))

    # ---------- 通用 ----------
    def pick(self, var):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm"), ("所有", "*.*")])
        if p:
            var.set(p)

    def pick_dir(self, var):
        p = filedialog.askdirectory()
        if p:
            var.set(p)

    def _add_row(self, parent):
        row = tk.Frame(parent)
        idx = len(self.map_rows)
        s = tk.StringVar(value="规格商家编码" if idx == 0 else "")
        t = tk.StringVar(value="规格商家编码" if idx == 0 else "")
        ttk.Label(row, text=f"映射{idx + 1}:").pack(side="left")
        ttk.Entry(row, textvariable=s, width=16).pack(side="left", padx=3)
        ttk.Label(row, text="→").pack(side="left")
        ttk.Entry(row, textvariable=t, width=16).pack(side="left", padx=3)
        row.pack(fill="x", padx=5, pady=2)
        self.map_rows.append((s, t))

    def log(self, box, txt):
        box.insert("end", txt + "\n")
        box.see("end")

    def run(self):
        src = self.src_var.get().strip().strip('"')
        tgt = self.tgt_var.get().strip().strip('"')
        if not src or not tgt:
            messagebox.showwarning("提示", "请先选择源表和目标表")
            return
        fill_map = [(s.get().strip(), t.get().strip())
                    for s, t in self.map_rows if s.get().strip() and t.get().strip()]
        if not fill_map:
            messagebox.showwarning("提示", "请至少填写一个回填映射")
            return
        self.run_btn.config(state="disabled")
        self.log(self.match_log, "▶ 开始核对...")
        try:
            matched, multi, notfound, out_path = run_match(
                src, tgt, self.sk.get().strip(), self.tk.get().strip(),
                fill_map, self.sku_col.get().strip() or None,
                self.rmk_col.get().strip() or None, self.skip_existing.get())
            self.log(self.match_log, f"✅ 匹配: {matched}, 多规格: {multi}, 未匹配: {len(notfound)}")
            self.log(self.match_log, f"   已另存(源文件未改): {os.path.basename(out_path)}")
            if notfound:
                self.log(self.match_log, f"\n⚠️ 未匹配 ({len(notfound)}):")
                for v in notfound[:50]:
                    self.log(self.match_log, f"   {v}")
                if len(notfound) > 50:
                    self.log(self.match_log, f"   ... 等 {len(notfound)} 个")
            messagebox.showinfo("成功", f"核对完成!\n匹配 {matched}\n多规格 {multi}\n未匹配 {len(notfound)}")
        except Exception as e:
            self.log(self.match_log, f"❌ 错误: {e}")
            messagebox.showerror("错误", str(e))
        finally:
            self.run_btn.config(state="normal")

    def run_img(self):
        src = self.img_src.get().strip().strip('"')
        if not src:
            messagebox.showwarning("提示", "请选择源文件夹")
            return
        q = max(1, min(100, self.img_q.get()))
        global QUALITY
        QUALITY = q
        if self.img_out.get().strip():
            out = self.img_out.get().strip().strip('"')
        else:
            out = src.rstrip('/\\') + "_jpg"
        self.img_btn.config(state="disabled")
        self.log(self.img_log, f"▶ 转换中... 源: {src} → {out}")
        try:
            total, converted, copied, failed, errors = run_imgconvert(src, out)
            self.log(self.img_log, f"✅ 完成: 总{total}, 转JPG{converted}, 复制{copied}, 失败{failed}")
            self.log(self.img_log, f"   输出: {out}")
            if errors:
                self.log(self.img_log, "⚠️ 失败清单:")
                for rel, e in errors[:10]:
                    self.log(self.img_log, f"   {rel}: {e}")
            messagebox.showinfo("成功", f"转换完成!\n总{total}\n转JPG {converted}\n复制 {copied}\n失败 {failed}")
        except Exception as e:
            self.log(self.img_log, f"❌ 错误: {e}")
            messagebox.showerror("错误", str(e))
        finally:
            self.img_btn.config(state="normal")


def main():
    root = tk.Tk()
    # 设置窗口/任务栏图标 (打包时 --add-data 附带 app.ico)
    try:
        if getattr(sys, "frozen", False):
            icon_path = os.path.join(sys._MEIPASS, "app.ico")
        else:
            icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "app.ico")
        if os.path.exists(icon_path):
            root.iconbitmap(icon_path)
    except Exception:
        pass
    try:
        ttk.Style().theme_use("clam")
    except Exception:
        pass
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
