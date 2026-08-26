#!/usr/bin/env python3
"""
电商工具 v6.0 - 表格核对 + 图片转换 + 开票生成
============================================
Tab1 表格核对: 用源表某字段去目标表检索, 回填字段 + 多规格标记
Tab2 图片转换: 批量图片转JPG(webp/png/heic等→jpg), 保留目录树
Tab3 开票生成: 数电发票批量开票模板生成(流水号自动+固定内容+手动/批量录入)

功能:
  - 任意检索键列(快递单号/订单号等)
  - 增量填充(已有数据不覆盖)
  - 多规格标记(≥2种商品→备注"多规格")
  - 输出另存加后缀(不修改源文件)
  - 图片批量转JPG, 保留目录结构
  - GUI(tkinter) + CLI 双界面
纯 tkinter 标准库 + openpyxl + Pillow
"""
import os
import re
import sys
import shutil
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from collections import defaultdict

import openpyxl

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

try:
    from invoice_gen import (generate_invoice_xlsx, parse_bulk_text,
                              DEFAULT_FIXED, INVOICE_TYPE_OPTIONS, find_template)
    INVOICE_AVAILABLE = True
except ImportError:
    INVOICE_AVAILABLE = False

try:
    from update import check_update, download_file, self_exe_name, self_asset_name, make_updater_bat, run_updater, CURRENT_VERSION
    UPDATE_AVAILABLE = True
except ImportError:
    UPDATE_AVAILABLE = False

try:
    from email_sender import (load_config, save_config, reset_config, send_email,
                              config_path, PRESETS)
    EMAIL_AVAILABLE = True
except ImportError:
    EMAIL_AVAILABLE = False



# ---------- 顶层弹窗 (所有messagebox置顶) ----------
_MBOX_ROOT = None

def _set_mbox_root(root):
    global _MBOX_ROOT
    _MBOX_ROOT = root

def _mtop(kind, *args, **kwargs):
    """带系统顶层的 messagebox 调用 (自动置顶, 避免被其他窗口遮挡)"""
    import tkinter.messagebox as _mb
    fn = getattr(_mb, kind)
    if _MBOX_ROOT is not None:
        try:
            _MBOX_ROOT.attributes('-topmost', True)
            _MBOX_ROOT.lift()
            kwargs['parent'] = _MBOX_ROOT
        except Exception:
            pass
    r = fn(*args, **kwargs)
    if _MBOX_ROOT is not None:
        try:
            _MBOX_ROOT.attributes('-topmost', False)
        except Exception:
            pass
    return r

# ================= 表格核对核心 =================
def _looks_like_number(s):
    """判断字符串是否为数字(金额/数量/订单号) — 允许逗号/连字符/空格"""
    try:
        s2 = str(s).replace(",", "").replace("，", "").replace("-", "").replace(" ", "")
        float(s2)
        return True
    except (ValueError, TypeError):
        return False


def find_header_row(ws, keywords):
    for r in range(1, min(6, ws.max_row + 1)):
        strs = [str(ws.cell(row=r, column=c + 1).value or '').strip()
                for c in range(ws.max_column)]
        if sum(1 for k in keywords if any(k.lower() in s.lower() for s in strs)) >= len(keywords):
            return r
    return None


def col_index_by_header(ws, header_row, header_name):
    for c in range(ws.max_column):
        if str(ws.cell(row=header_row, column=c + 1).value or '').strip() == header_name:
            return c
    return None


def _tolerant_load(path, data_only=False):
    """加载xlsx, 遇 Fill 样式错误自动修复 styles.xml 重载(快麦等导出文件兼容)"""
    try:
        return openpyxl.load_workbook(path, data_only=data_only)
    except (TypeError, KeyError, IndexError) as e:
        if "Fill" not in str(e) and "fill" not in str(e):
            raise
        # 修复 styles.xml 的 fills 段: 全部替换为默认 patternFill
        import zipfile, tempfile, re as _re
        tmp = tempfile.mktemp(suffix="_fix.xlsx")
        try:
            with zipfile.ZipFile(path) as zin:
                with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                    for n in zin.namelist():
                        data = zin.read(n)
                        if n == "xl/styles.xml":
                            xml = data.decode("utf-8", errors="ignore")
                            m = _re.search(r'<fills count="(\d+)">.*?</fills>', xml, _re.S)
                            if m:
                                cnt = int(m.group(1))
                                if cnt <= 2:
                                    new_fills = f'<fills count="{cnt}">' + "<fill><patternFill patternType=\"none\"/></fill>" * cnt + "</fills>"
                                else:
                                    new_fills = ('<fills count="' + str(cnt) + '">'
                                                 '<fill><patternFill patternType="none"/></fill>'
                                                 '<fill><patternFill patternType="gray125"/></fill>'
                                                 + '<fill><patternFill patternType="none"/></fill>' * (cnt - 2)
                                                 + '</fills>')
                                xml = xml[:m.start()] + new_fills + xml[m.end():]
                            data = xml.encode("utf-8")
                        zout.writestr(n, data)
            return openpyxl.load_workbook(tmp, data_only=data_only)
        finally:
            try:
                os.remove(tmp)
            except Exception:
                pass


def run_match(src_path, tgt_path, src_key, tgt_key, fill_map, sku_col, remark_col,
              skip_existing, progress_cb=None):
    twb = _tolerant_load(tgt_path, data_only=True)
    tws = twb[twb.sheetnames[0]]
    thr = find_header_row(tws, [tgt_key])
    if thr is None:
        raise ValueError(f"目标表找不到表头(含 '{tgt_key}')")
    key_col = col_index_by_header(tws, thr, tgt_key)
    if key_col is None:
        raise ValueError(f"目标表没有列 '{tgt_key}'")
    tgt_cols = {}
    for sc, tc in fill_map:
        ci = col_index_by_header(tws, thr, tc)
        if ci is not None:
            tgt_cols.setdefault(sc, []).append(ci)
    sku_idx = col_index_by_header(tws, thr, sku_col) if sku_col else None

    idx = defaultdict(list)
    for r in tws.iter_rows(min_row=thr + 1, values_only=True):
        k = r[key_col]
        if not k:
            continue
        for part in str(k).replace('，', ',').replace(' ', '').split(','):
            if part.strip():
                idx[part.strip()].append(r)

    swb = _tolerant_load(src_path)
    sws = swb[swb.sheetnames[0]]
    shr = find_header_row(sws, [src_key])
    if shr is None:
        raise ValueError(f"源表找不到表头(含 '{src_key}')")
    src_key_col = col_index_by_header(sws, shr, src_key)
    src_fill_cols = {}
    for sc, tc in fill_map:
        ci = col_index_by_header(sws, shr, sc)
        if ci is not None:
            src_fill_cols[sc] = ci
    remark_col_idx = col_index_by_header(sws, shr, remark_col) if remark_col else None

    matched = 0
    multi = 0
    notfound = []
    total = 0
    for row in sws.iter_rows(min_row=shr + 1):
        k = row[src_key_col].value
        if not k:
            continue
        total += 1
        k = str(k).strip()
        hits = idx.get(k)
        if hits:
            matched += 1
            for sc, tc in fill_map:
                sci = src_fill_cols.get(sc)
                tci_list = tgt_cols.get(sc)
                if sci is not None and tci_list and not row[sci].value:
                    row[sci].value = hits[0][tci_list[0]]
            if remark_col_idx is not None and not row[remark_col_idx].value:
                if sku_idx is not None and len(hits) >= 2:
                    uniq = {str(h[sku_idx] or '').strip() for h in hits if h[sku_idx]}
                    if len(uniq) >= 2:
                        row[remark_col_idx].value = "多规格"
                        multi += 1
        else:
            notfound.append(k)
        if progress_cb and total % 10 == 0:
            progress_cb(total)

    base, ext = os.path.splitext(src_path)
    out_path = f"{base}_结果{ext}"
    swb.save(out_path)
    return matched, multi, notfound, out_path


# ================= 图片转换核心 =================
IMG_EXTS = {'.webp', '.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff', '.heic', '.avif', '.jfif'}
QUALITY = 92


def convert_one(src, dst):
    img = Image.open(src)
    if img.mode in ('RGBA', 'LA', 'P', 'PA'):
        rgba = img.convert('RGBA')
        bg = Image.new('RGB', rgba.size, (255, 255, 255))
        bg.paste(rgba, mask=rgba.split()[-1])
        img = bg
    elif img.mode != 'RGB':
        img = img.convert('RGB')
    img.save(dst, 'JPEG', quality=QUALITY)


def run_imgconvert(src_root, out_root, progress_cb=None):
    """批量转JPG, 保留目录树. 返回 (total, converted, copied, failed, errors)"""
    if not PIL_AVAILABLE:
        raise ValueError("缺少 Pillow 库, 无法转换图片. 请运行: pip install pillow")
    if not os.path.isdir(src_root):
        raise ValueError(f"路径不存在: {src_root}")
    os.makedirs(out_root, exist_ok=True)
    total = converted = copied = failed = 0
    errors = []
    abs_out = os.path.abspath(out_root)
    count = 0
    for root, dirs, files in os.walk(src_root):
        if os.path.abspath(root).startswith(abs_out):
            continue
        for f in files:
            src = os.path.join(root, f)
            rel = os.path.relpath(src, src_root)
            total += 1
            try:
                if os.path.splitext(f)[1].lower() in IMG_EXTS:
                    base = os.path.splitext(f)[0]
                    dst = os.path.join(out_root, os.path.dirname(rel), base + ".jpg")
                    os.makedirs(os.path.dirname(dst), exist_ok=True)
                    convert_one(src, dst)
                    converted += 1
                else:
                    dst = os.path.join(out_root, rel)
                    os.makedirs(os.path.dirname(dst), exist_ok=True)
                    shutil.copy2(src, dst)
                    copied += 1
            except Exception as e:
                failed += 1
                errors.append((rel, str(e)))
            count += 1
            if progress_cb and count % 20 == 0:
                progress_cb(count, converted, failed)
    return total, converted, copied, failed, errors


# ================= GUI =================
class App:
    def __init__(self, root):
        self.root = root
        # 版本号从 update.py 动态读取(发布只需改一处)
        try:
            root.title("电商工具 " + CURRENT_VERSION)
        except Exception:
            root.title("电商工具 v6.4.0-Flex (灵活开票版) BY 大萝北拔萝卜")
        root.geometry("780x720")
        root.minsize(700, 640)

        # 公共底部状态栏(所有Tab可见) — 必须先pack(占bottom), 否则被Notebook挤掉
        status = ttk.Frame(root)
        status.pack(side="bottom", fill="x", padx=10, pady=(0, 5))
        try:
            _cur_ver = CURRENT_VERSION
        except Exception:
            _cur_ver = "?"
        self.status_ver = tk.StringVar(value=f"v{_cur_ver}")
        # 版本号(点击可检查更新)
        ver_lbl = ttk.Label(status, textvariable=self.status_ver, foreground="#666")
        ver_lbl.pack(side="left")
        ver_lbl.bind("<Button-1>", lambda e: self.check_update_btn())
        # 全局状态信息(所有Tab可见)
        self.status_msg = tk.StringVar(value="")
        ttk.Label(status, textvariable=self.status_msg, foreground="#2a7", width=36).pack(side="left", padx=10)
        # 检查更新按钮(保存引用以便改提示文字)
        self.update_btn = ttk.Button(status, text="🔍 检查更新 (v)", command=self.check_update_btn)
        self.update_btn.pack(side="right")
        # 下载进度条(更新时显示)
        self.dl_progress = ttk.Progressbar(status, mode="determinate", length=160)
        self.dl_progress.pack(side="right", padx=8)
        self.dl_progress.pack_forget()  # 默认隐藏, 下载时显示
        self.dl_pct = tk.StringVar(value="")
        ttk.Label(status, textvariable=self.dl_pct, width=8).pack(side="right")

        self.nb = ttk.Notebook(root)
        self.nb.pack(fill="both", expand=True, padx=6, pady=6)

        self.build_match_tab()
        self.build_img_tab()
        self.build_invoice_tab()
        self.build_email_tab()

    # ---------- Tab1 表格核对 ----------
    def build_match_tab(self):
        tab = ttk.Frame(self.nb)
        self.nb.add(tab, text="① 表格核对")

        frm = ttk.LabelFrame(tab, text="1. 选择文件")
        frm.pack(fill="x", padx=10, pady=5)
        self.src_var = tk.StringVar()
        self.tgt_var = tk.StringVar()
        ttk.Label(frm, text="源表(被填的):").grid(row=0, column=0, sticky="w", padx=5, pady=3)
        ttk.Entry(frm, textvariable=self.src_var, width=45).grid(row=0, column=1, padx=5)
        ttk.Button(frm, text="浏览", command=lambda: self.pick(self.src_var)).grid(row=0, column=2)
        ttk.Label(frm, text="目标表(数据源):").grid(row=1, column=0, sticky="w", padx=5)
        ttk.Entry(frm, textvariable=self.tgt_var, width=45).grid(row=1, column=1, padx=5)
        ttk.Button(frm, text="浏览", command=lambda: self.pick(self.tgt_var)).grid(row=1, column=2)

        frm2 = ttk.LabelFrame(tab, text="2. 检索键列 (表头名)")
        frm2.pack(fill="x", padx=10, pady=5)
        self.sk = tk.StringVar(value="快递单号")
        self.tk = tk.StringVar(value="快递单号")
        ttk.Label(frm2, text="源表键列:").grid(row=0, column=0, padx=5)
        ttk.Entry(frm2, textvariable=self.sk, width=20).grid(row=0, column=1, padx=5)
        ttk.Label(frm2, text="目标表键列:").grid(row=0, column=2, padx=5)
        ttk.Entry(frm2, textvariable=self.tk, width=20).grid(row=0, column=3, padx=5)

        frm3 = ttk.LabelFrame(tab, text="3. 回填映射 (源列 → 目标列)")
        frm3.pack(fill="x", padx=10, pady=5)
        self.map_rows = []
        self._add_row(frm3)
        self.add_btn = ttk.Button(frm3, text="+ 添加映射", command=lambda: self._add_row(frm3))
        self.add_btn.pack(anchor="w", padx=10, pady=3)

        frm4 = ttk.LabelFrame(tab, text="4. 多规格标记 (可空)")
        frm4.pack(fill="x", padx=10, pady=5)
        self.sku_col = tk.StringVar()
        self.rmk_col = tk.StringVar()
        ttk.Label(frm4, text="多规格判断列:").grid(row=0, column=0, padx=5)
        ttk.Entry(frm4, textvariable=self.sku_col, width=18).grid(row=0, column=1, padx=5)
        ttk.Label(frm4, text="备注列:").grid(row=0, column=2, padx=5)
        ttk.Entry(frm4, textvariable=self.rmk_col, width=18).grid(row=0, column=3, padx=5)

        frm5 = ttk.Frame(tab)
        frm5.pack(fill="x", padx=10, pady=5)
        self.skip_existing = tk.BooleanVar(value=True)
        ttk.Checkbutton(frm5, text="已有数据不覆盖", variable=self.skip_existing).pack(side="left", padx=5)
        self.run_btn = ttk.Button(frm5, text="▶ 开始核对", command=self.run)
        self.run_btn.pack(side="right", padx=5)

        self.match_log = scrolledtext.ScrolledText(tab, height=12, font=("Consolas", 9))
        self.match_log.pack(fill="both", expand=True, padx=10, pady=5)

    # ---------- Tab2 图片转换 ----------
    def build_img_tab(self):
        tab = ttk.Frame(self.nb)
        self.nb.add(tab, text="② 图片转JPG")

        frm = ttk.LabelFrame(tab, text="1. 选择图片文件夹")
        frm.pack(fill="x", padx=10, pady=5)
        self.img_src = tk.StringVar()
        self.img_out = tk.StringVar()
        ttk.Label(frm, text="源文件夹:").grid(row=0, column=0, sticky="w", padx=5, pady=3)
        ttk.Entry(frm, textvariable=self.img_src, width=45).grid(row=0, column=1, padx=5)
        ttk.Button(frm, text="浏览", command=lambda: self.pick_dir(self.img_src)).grid(row=0, column=2)
        ttk.Label(frm, text="输出路径(可空=自动):").grid(row=1, column=0, sticky="w", padx=5)
        ttk.Entry(frm, textvariable=self.img_out, width=45).grid(row=1, column=1, padx=5)
        ttk.Button(frm, text="浏览", command=lambda: self.pick_dir(self.img_out)).grid(row=1, column=2)

        frm2 = ttk.Frame(tab)
        frm2.pack(fill="x", padx=10, pady=5)
        ttk.Label(frm2, text="质量 (1-100):").pack(side="left")
        self.img_q = tk.IntVar(value=92)
        ttk.Spinbox(frm2, from_=1, to=100, textvariable=self.img_q, width=5).pack(side="left", padx=5)
        self.img_btn = ttk.Button(frm2, text="▶ 开始转换", command=self.run_img)
        self.img_btn.pack(side="right", padx=5)

        self.img_log = scrolledtext.ScrolledText(tab, height=12, font=("Consolas", 9))
        self.img_log.pack(fill="both", expand=True, padx=10, pady=5)

    # ---------- Tab3 开票生成 ----------
    def build_invoice_tab(self):
        tab = ttk.Frame(self.nb)
        self.nb.add(tab, text="③ 开票生成")

        # 0. 固定内容配置 (多行grid, 不挤宽度, 窗口可扩大)
        frm0 = ttk.LabelFrame(tab, text="开票内容 (每次开票填写: 项目名称/税收编码/单位/税率)")
        frm0.pack(fill="x", padx=10, pady=5)
        self.inv_type = tk.StringVar(value=DEFAULT_FIXED["invoice_type"])
        self.inv_taxinc = tk.StringVar(value=DEFAULT_FIXED["tax_included"])
        self.inv_item = tk.StringVar(value=DEFAULT_FIXED["item_name"])
        self.inv_code = tk.StringVar(value=DEFAULT_FIXED["tax_code"])
        self.inv_unit = tk.StringVar(value=DEFAULT_FIXED["unit"])
        self.inv_rate = tk.StringVar(value=DEFAULT_FIXED["tax_rate"])
        # grid 布局: 每行3组 (标签+控件)
        row = 0
        ttk.Label(frm0, text="发票类型:").grid(row=row, column=0, sticky="e", padx=(8, 2), pady=3)
        ttk.Combobox(frm0, textvariable=self.inv_type, values=INVOICE_TYPE_OPTIONS,
                     width=12, state="readonly").grid(row=row, column=1, sticky="w", padx=(0, 10), pady=3)
        ttk.Label(frm0, text="是否含税:").grid(row=row, column=2, sticky="e", padx=(8, 2), pady=3)
        ttk.Combobox(frm0, textvariable=self.inv_taxinc, values=["是", "否"],
                     width=4, state="readonly").grid(row=row, column=3, sticky="w", padx=(0, 10), pady=3)
        ttk.Label(frm0, text="项目名称:").grid(row=row, column=4, sticky="e", padx=(8, 2), pady=3)
        ttk.Entry(frm0, textvariable=self.inv_item, width=10).grid(row=row, column=5, sticky="w", padx=(0, 10), pady=3)
        row += 1
        ttk.Label(frm0, text="税收编码:").grid(row=row, column=0, sticky="e", padx=(8, 2), pady=3)
        ttk.Entry(frm0, textvariable=self.inv_code, width=22).grid(row=row, column=1, sticky="w", padx=(0, 10), pady=3)
        ttk.Label(frm0, text="单位:").grid(row=row, column=2, sticky="e", padx=(8, 2), pady=3)
        ttk.Entry(frm0, textvariable=self.inv_unit, width=4).grid(row=row, column=3, sticky="w", padx=(0, 10), pady=3)
        ttk.Label(frm0, text="税率:").grid(row=row, column=4, sticky="e", padx=(8, 2), pady=3)
        ttk.Entry(frm0, textvariable=self.inv_rate, width=6).grid(row=row, column=5, sticky="w", padx=(0, 10), pady=3)
        row += 1
        self.inv_tpl = tk.StringVar(value="")
        ttk.Label(frm0, text="开票模板:").grid(row=row, column=0, sticky="e", padx=(8, 2), pady=3)
        ttk.Entry(frm0, textvariable=self.inv_tpl, width=38).grid(row=row, column=1, columnspan=3, sticky="we", padx=(0, 6), pady=3)
        ttk.Button(frm0, text="浏览…", command=lambda: self.pick(self.inv_tpl)).grid(row=row, column=4, sticky="w", padx=(0, 8), pady=3)
        frm0.columnconfigure(1, weight=1)
        frm0.columnconfigure(3, weight=1)

        # 1. 可编辑发票表格 (每行一个输入框 = 每行一张发票)
        frm2 = ttk.LabelFrame(tab, text="发票录入表 — 双击单元格填写; 点选某列后Ctrl+V只粘该列; 整块复制自动对齐 (每行一张发票)")
        frm2.pack(fill="both", expand=True, padx=10, pady=5)
        cols = ("serial", "invtype", "taxinc", "buyer", "taxid", "natural", "qty", "amount", "remark")
        self.inv_tree = ttk.Treeview(frm2, columns=cols, show="headings", height=12)
        headers = {"serial": "流水号", "invtype": "发票类型", "taxinc": "含税",
                   "buyer": "购买方名称*", "taxid": "纳税人识别号",
                   "natural": "自然人", "qty": "数量", "amount": "金额*", "remark": "备注"}
        widths = {"serial": 48, "invtype": 90, "taxinc": 48, "buyer": 150, "taxid": 150,
                  "natural": 50, "qty": 50, "amount": 80, "remark": 110}
        for c in cols:
            self.inv_tree.heading(c, text=headers[c])
            self.inv_tree.column(c, width=widths[c], anchor="w")
        vsb = ttk.Scrollbar(frm2, orient="vertical", command=self.inv_tree.yview)
        self.inv_tree.configure(yscrollcommand=vsb.set)
        # 网格线样式(横线+竖线分隔 + 斑马纹)
        try:
            st = ttk.Style(self.root)
            st.configure("Inv.Treeview", rowheight=26,
                         bordercolor="#555555", lightcolor="#555555", darkcolor="#555555",
                         fieldbackground="white", background="white")
            st.configure("Inv.Treeview.Heading", bordercolor="#888888",
                         lightcolor="#888888", darkcolor="#888888",
                         font=("Microsoft YaHei", 9, "bold"))
            st.map("Inv.Treeview",
                   background=[("selected", "#d0e4f7")],
                   foreground=[("selected", "#000000")])
            self.inv_tree.configure(style="Inv.Treeview")
            # 斑马纹: 交替行背景增强横线分隔感
            def _zebra():
                try:
                    for i, iid in enumerate(self.inv_tree.get_children()):
                        tag = "odd" if i % 2 else "even"
                        self.inv_tree.item(iid, tags=(tag,))
                except Exception:
                    pass
            self.inv_tree.tag_configure("odd", background="#f7f7f7")
            self.inv_tree.tag_configure("even", background="#ffffff")
            self._inv_zebra = _zebra
        except Exception:
            self._inv_zebra = None
        self.inv_tree.pack(side="left", fill="both", expand=True, padx=(5, 0), pady=5)
        vsb.pack(side="right", fill="y", pady=5)
        self.invoices = []  # 数据模型: [{invoice_type,tax_included,buyer,tax_id,is_natural,qty,amount,remark}]
        self._inv_editing = None  # 当前编辑状态 (entry, row_id, col_idx)
        self._inv_click_col = None  # 粘贴起点列(点选记录) 0=名称 1=税号 2=自然人 3=数量 4=金额 5=备注
        self._inv_click_row = None  # 粘贴起点行索引
        self._inv_click_row_id = None  # 粘贴起点行Treeview id

        # 双击编辑 + Ctrl+V 粘贴 + 单击记录列 绑定
        self.inv_tree.bind("<Double-1>", self.inv_cell_double_click)
        self.inv_tree.bind("<Control-v>", self.inv_paste)
        self.inv_tree.bind("<Control-V>", self.inv_paste)
        self.inv_tree.bind("<Button-1>", self.inv_click_col)

        # 2. 操作按钮
        frm3 = ttk.Frame(tab)
        frm3.pack(fill="x", padx=10, pady=5)
        ttk.Button(frm3, text="＋ 新增一行", command=self.invoice_add_row).pack(side="left", padx=5)
        ttk.Button(frm3, text="📥 导入明细文档", command=self.invoice_import_detail).pack(side="left", padx=5)
        ttk.Button(frm3, text="删除选中行", command=self.invoice_del_sel).pack(side="left", padx=5)
        ttk.Button(frm3, text="清空选中列", command=self.invoice_clear_col).pack(side="left", padx=5)
        ttk.Button(frm3, text="清空", command=self.invoice_clear).pack(side="left", padx=5)
        ttk.Label(frm3, text="单列粘贴→:").pack(side="left", padx=(12, 2))
        self.inv_paste_col = tk.StringVar(value="自动")
        paste_opts = ["自动", "发票类型", "含税", "名称", "税号", "自然人", "数量", "金额", "备注"]
        ttk.Combobox(frm3, textvariable=self.inv_paste_col, values=paste_opts,
                     width=5, state="readonly").pack(side="left", padx=2)
        ttk.Button(frm3, text="▶ 生成开票xlsx", command=self.invoice_generate).pack(side="right", padx=5)

        self.inv_log = scrolledtext.ScrolledText(tab, height=5, font=("Consolas", 9))
        self.inv_log.pack(fill="both", expand=True, padx=10, pady=5)
        ttk.Label(tab, text="BY 大萝北拔萝卜", foreground="#888").pack(anchor="e", padx=12, pady=(0, 4))

        if not INVOICE_AVAILABLE:
            self.log(self.inv_log, "⚠️ 开票模块未加载(invoice_gen.py缺失), 请联系管理员")

    # ---------- Tab4 邮箱发送 ----------
    def build_email_tab(self):
        tab = ttk.Frame(self.nb)
        self.nb.add(tab, text="④ 邮箱发送")

        if not EMAIL_AVAILABLE:
            ttk.Label(tab, text="⚠️ 邮件模块未加载(email_sender.py缺失)").pack(pady=20)
            return

        # 状态栏: 配置状态
        frm0 = ttk.Frame(tab)
        frm0.pack(fill="x", padx=10, pady=5)
        cfg = load_config()
        self.email_cfg_state = tk.StringVar(
            value=f"发送邮箱: {cfg['sender_email']}" if cfg else "未配置 — 首次使用请先设置")
        ttk.Label(frm0, textvariable=self.email_cfg_state, foreground="#c55"
                  if not cfg else "#383").pack(side="left")
        ttk.Button(frm0, text="⚙ 设置邮箱", command=self.email_config_dlg).pack(side="left", padx=10)
        ttk.Button(frm0, text="重置配置", command=self.email_reset_cfg).pack(side="left")

        # 邮件内容区
        frm1 = ttk.LabelFrame(tab, text="邮件内容")
        frm1.pack(fill="both", expand=True, padx=10, pady=5)
        row = 0
        ttk.Label(frm1, text="收件人:").grid(row=row, column=0, sticky="e", padx=6, pady=6)
        self.email_to = tk.StringVar()
        ttk.Entry(frm1, textvariable=self.email_to, width=60).grid(row=row, column=1, sticky="we", padx=6)
        row += 1
        ttk.Label(frm1, text="主题:").grid(row=row, column=0, sticky="e", padx=6, pady=6)
        self.email_subj = tk.StringVar()
        ttk.Entry(frm1, textvariable=self.email_subj, width=60).grid(row=row, column=1, sticky="we", padx=6)
        row += 1
        ttk.Label(frm1, text="正文:").grid(row=row, column=0, sticky="ne", padx=6, pady=6)
        self.email_body = scrolledtext.ScrolledText(frm1, height=6, font=("Microsoft YaHei", 9))
        self.email_body.grid(row=row, column=1, sticky="nsew", padx=6, pady=6)
        row += 1
        ttk.Label(frm1, text="附件:").grid(row=row, column=0, sticky="e", padx=6, pady=6)
        self.email_attach = tk.StringVar()
        ttk.Entry(frm1, textvariable=self.email_attach, width=60).grid(row=row, column=1, sticky="we", padx=6)
        ttk.Button(frm1, text="选择文件…", command=self.email_pick_attach).grid(row=row, column=2, padx=6)
        frm1.columnconfigure(1, weight=1)

        # 发送/日志
        frm2 = ttk.Frame(tab)
        frm2.pack(fill="x", padx=10, pady=5)
        ttk.Button(frm2, text="📨 发送邮件", command=self.email_send).pack(side="left", padx=5)
        self.email_log = scrolledtext.ScrolledText(tab, height=6, font=("Consolas", 9))
        self.email_log.pack(fill="both", expand=True, padx=10, pady=5)
        ttk.Label(tab, text="BY 大萝北拔萝卜", foreground="#888").pack(anchor="e", padx=12, pady=(0, 4))
        self.log(self.email_log, "提示: 填好收件人/主题/正文/附件, 点『发送邮件』")

    def email_config_dlg(self):
        """设置邮箱弹窗 (SMTP 配置), 敏感字段加密保存"""
        dlg = tk.Toplevel(self.root)
        dlg.title("设置发送邮箱 (SMTP)")
        dlg.geometry("480x330")
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.attributes("-topmost", True)  # 系统顶层

        cfg = load_config() or {}
        frame = ttk.Frame(dlg, padding=14)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="预设:").grid(row=0, column=0, sticky="e", padx=6, pady=6)
        presets = tk.StringVar()
        combo = ttk.Combobox(frame, textvariable=presets, values=list(PRESETS.keys()),
                             width=20, state="readonly")
        combo.grid(row=0, column=1, sticky="w", padx=6)
        def _apply_preset(*_):
            p = PRESETS.get(presets.get())
            if p:
                sv_server.set(p["smtp_server"])
                sv_port.set(str(p["smtp_port"]))
        combo.bind("<<ComboboxSelected>>", _apply_preset)

        ttk.Label(frame, text="SMTP服务器:").grid(row=1, column=0, sticky="e", padx=6, pady=6)
        sv_server = tk.StringVar(value=cfg.get("smtp_server", "smtp.qq.com"))
        ttk.Entry(frame, textvariable=sv_server, width=35).grid(row=1, column=1, sticky="w", padx=6)
        ttk.Label(frame, text="端口:").grid(row=1, column=2, sticky="e", padx=4)
        sv_port = tk.StringVar(value=str(cfg.get("smtp_port", 465)))
        ttk.Entry(frame, textvariable=sv_port, width=8).grid(row=1, column=3, sticky="w", padx=6)

        ttk.Label(frame, text="发件邮箱:").grid(row=2, column=0, sticky="e", padx=6, pady=6)
        sv_sender = tk.StringVar(value=cfg.get("sender_email", ""))
        ttk.Entry(frame, textvariable=sv_sender, width=35).grid(row=2, column=1, columnspan=3, sticky="w", padx=6)

        ttk.Label(frame, text="授权码/密码:").grid(row=3, column=0, sticky="e", padx=6, pady=6)
        sv_auth = tk.StringVar(value=cfg.get("auth_code", ""))
        ttk.Entry(frame, textvariable=sv_auth, width=35, show="*").grid(row=3, column=1, columnspan=3, sticky="w", padx=6)

        ttk.Label(frame, text="提示: QQ/163邮箱需用『授权码』而非登录密码\n发件人将显示为邮箱地址(纯地址)", foreground="#888").grid(
            row=4, column=0, columnspan=4, sticky="w", padx=6, pady=4)

        def _ok():
            sender = sv_sender.get().strip()
            auth = sv_auth.get().strip()
            server = sv_server.get().strip()
            try:
                port = int(sv_port.get().strip() or 465)
            except ValueError:
                port = 465
            if not sender or not auth or not server:
                _mtop('showerror', "错误", "SMTP服务器/发件邮箱/授权码 不能为空")
                return
            new_cfg = {"smtp_server": server, "smtp_port": port, "use_ssl": True,
                       "sender_email": sender, "auth_code": auth}
            path = save_config(new_cfg)
            self.email_cfg_state.set(f"发送邮箱: {sender}")
            self.log(self.email_log, f"✅ 邮箱配置已保存: {path}")
            _mtop('showinfo', "成功", "邮箱配置已保存(已加密)")
            dlg.destroy()

        btns = ttk.Frame(dlg)
        btns.pack(pady=8)
        ttk.Button(btns, text="保存", command=_ok).pack(side="left", padx=10)
        ttk.Button(btns, text="取消", command=dlg.destroy).pack(side="left", padx=10)

    def email_reset_cfg(self):
        if _mtop('askyesno', "确认", "删除已保存的邮箱配置?"):
            reset_config()
            self.email_cfg_state.set("未配置 — 首次使用请先设置")
            self.log(self.email_log, "🗑 邮箱配置已删除")

    def email_pick_attach(self):
        p = filedialog.askopenfilename(filetypes=[("所有文件", "*.*")])
        if p:
            self.email_attach.set(p)

    def email_send(self):
        cfg = load_config()
        if not cfg:
            _mtop('showinfo', "提示", "请先设置邮箱(SMTP配置)")
            self.email_config_dlg()
            return
        to = self.email_to.get().strip()
        subj = self.email_subj.get().strip()
        body = self.email_body.get("1.0", "end").strip()
        attach = self.email_attach.get().strip().strip('"')
        if not to:
            _mtop('showwarning', "提示", "收件人为空")
            return
        # 后台线程发送, 避免阻塞UI(未响应)
        self.log(self.email_log, f"📨 发送中 → {to} ... (请稍候)")
        import threading

        def _work():
            ok, msg = send_email(cfg, to, subj, body, attachment=attach or None)
            self.root.after(0, lambda: self._email_done(ok, msg, to))

        t = threading.Thread(target=_work, daemon=True)
        t.start()

    def _email_done(self, ok, msg, to):
        if ok:
            self.log(self.email_log, f"✅ {msg} → {to}")
            _mtop('showinfo', "成功", "邮件已发送!")
        else:
            self.log(self.email_log, f"❌ 发送失败: {msg}")
            _mtop('showerror', "发送失败", str(msg))

    # ---------- 可编辑表格支持 ----------
    INV_COLS = ("serial", "invtype", "taxinc", "buyer", "taxid", "natural", "qty", "amount", "remark")

    def _inv_close_editor(self):
        """关闭当前编辑器"""
        if self._inv_editing:
            entry, row_id, col_idx = self._inv_editing
            try:
                entry.destroy()
            except Exception:
                pass
            self._inv_editing = None

    def inv_click_col(self, event):
        """单击单元格 → 记录粘贴起点(行+列)。下次Ctrl+V从该格开始, 粘完自动恢复自动"""
        # 单击时关闭残留编辑器(下拉/输入框)
        self._inv_close_editor()
        region = self.inv_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col_id = self.inv_tree.identify_column(event.x)
        row_id = self.inv_tree.identify_row(event.y)
        if not col_id:
            return
        col_idx = int(col_id.replace("#", "")) - 1
        if col_idx == 0:  # 流水号列忽略
            return
        # 列映射: #2=发票类型(0) #3=含税(1) #4=名称(2) #5=税号(3) #6=自然人(4) #7=数量(5) #8=金额(6) #9=备注(7)
        map_col = {"#2": 0, "#3": 1, "#4": 2, "#5": 3, "#6": 4, "#7": 5, "#8": 6, "#9": 7}
        col_key = map_col.get(col_id, None)
        if col_key is None:
            return
        # 记录 (行索引, 列key) 作为粘贴起点
        row_idx = self.inv_tree.index(row_id) if row_id else 0
        self._inv_click_col = col_key
        self._inv_click_row = row_idx
        self._inv_click_row_id = row_id

    def inv_cell_double_click(self, event):
        """双击单元格 → 就地编辑 (流水号列只读)"""
        region = self.inv_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        row_id = self.inv_tree.identify_row(event.y)
        col_id = self.inv_tree.identify_column(event.x)
        if not row_id or not col_id:
            return
        col_idx = int(col_id.replace("#", "")) - 1
        if col_idx < 0 or col_idx >= len(self.INV_COLS):
            return
        if col_idx == 0:  # 流水号只读
            return

        self._inv_close_editor()

        x, y, w, h = self.inv_tree.bbox(row_id, col_id)
        if x is None:
            return
        cur_val = self.inv_tree.set(row_id, col_id)

        # 发票类型/含税列 → 下拉选择(替代手动输入)
        if col_id in ("#2", "#3"):
            opts = ["普通发票", "增值税专用发票"] if col_id == "#2" else ["是", "否"]
            combo = ttk.Combobox(self.inv_tree, values=opts, state="readonly", width=w // 9)
            combo.place(x=x, y=y, width=w, height=h)
            combo.set(cur_val or opts[0])

            def save_combo(_=None):
                val = combo.get().strip()
                self._inv_close_editor()
                self.inv_tree.set(row_id, col_id, val)
                idx = self.inv_tree.index(row_id)
                key2 = ("serial", "invoice_type", "tax_included", "buyer", "tax_id", "is_natural", "qty", "amount", "remark")[col_idx]
                if idx < len(self.invoices) and key2 not in ("serial",):
                    self.invoices[idx][key2] = val
                self.invoice_refresh_tree()

            def cancel_combo(_=None):
                self._inv_close_editor()

            # 只绑选中保存; 不绑FocusOut(点下拉箭头展开列表会误触发导致关闭)
            combo.bind("<<ComboboxSelected>>", save_combo)
            combo.bind("<Escape>", cancel_combo)
            self._inv_editing = (combo, row_id, col_idx)
            combo.focus_set()
            return

        entry = tk.Entry(self.inv_tree, font=("Microsoft YaHei", 9))
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, cur_val or "")
        entry.focus_set()
        entry.select_range(0, "end")

        def save(_=None):
            val = entry.get().strip()
            self._inv_close_editor()
            self.inv_tree.set(row_id, col_id, val)
            # 同步数据模型
            idx = self.inv_tree.index(row_id)
            key = ("serial", "invoice_type", "tax_included", "buyer", "tax_id", "is_natural", "qty", "amount", "remark")[col_idx]
            if 0 <= idx < len(self.invoices):
                if key == "is_natural":
                    self.invoices[idx][key] = "是" if val == "是" else (val or "")
                else:
                    self.invoices[idx][key] = val

        entry.bind("<Return>", save)
        entry.bind("<FocusOut>", save)
        entry.bind("<Escape>", lambda e: self._inv_close_editor())
        self._inv_editing = (entry, row_id, col_idx)

    def inv_paste(self, event=None):
        """Ctrl+V 粘贴: 支持 Excel 复制的多行多列 (Tab=列, 换行=行)
        仅粘贴非流水号列; 粘贴数据超过现有行则自动扩展"""
        try:
            data = self.root.clipboard_get()
        except Exception:
            return "break"
        if not data:
            return "break"
        rows = [r.split("\t") for r in data.replace("\r\n", "\n").replace("\r", "\n").split("\n")]
        # 清洗: 去掉每个单元格首尾空白/\t/引号/多余符号; 并裁掉每行末尾空单元格
        cleaned = []
        for r in rows:
            rr = [c.strip().strip('"\'').replace("\u00a0", " ").strip() for c in r]
            while rr and rr[-1] == "":
                rr.pop()  # 去掉尾部空列(CSV的\t尾巴)
            if any(c for c in rr):
                cleaned.append(rr)
        rows = cleaned
        if not rows:
            return "break"

        self._inv_close_editor()
        # 起始行: 有点选的格子则从该格行开始, 否则选中行, 否则第0行
        if getattr(self, "_inv_click_row", None) is not None:
            start_row = self._inv_click_row
        else:
            sel = self.inv_tree.selection()
            start_row = self.inv_tree.index(sel[0]) if sel else 0

        # 判断起始列:
        #  0) 点选的列 > 手动指定 > 自动识别
        #  1) 两列纯数字(金额+数量场景): 有小数列→金额, 整数列→数量
        #  2) 多列(≥3): 从名称列开始对齐
        #  3) 单列纯数字: 全小数→金额; 全整数且≤4位→数量; 长数字(>4位, 如订单号)→备注
        #  4) 单列文字/"是/否": 名称列
        ncols_data = max(len(r) for r in rows)
        key_cols = ("invoice_type", "tax_included", "buyer", "tax_id", "is_natural", "qty", "amount", "remark")
        manual = {"发票类型": 0, "含税": 1, "名称": 2, "税号": 3, "自然人": 4,
                  "数量": 5, "金额": 6, "备注": 7}.get(
            self.inv_paste_col.get().strip(), -1)

        if self._inv_click_col is not None:
            start_key = self._inv_click_col  # 点选列优先: 只粘到点选的列
        elif manual >= 0:
            start_key = manual  # 手动指定优先
        elif ncols_data == 2 and all(
                all(_looks_like_number(c) for c in r if c.strip()) for r in rows):
            # 两列纯数字: 检测每列含小数情况
            col_has_decimal = [False, False]
            for r in rows:
                for j in (0, 1):
                    if j < len(r) and r[j].strip():
                        if "." in r[j] or "．" in r[j]:
                            col_has_decimal[j] = True
            # 有小数列视为金额(下标6), 另一列视为数量(下标5)
            if col_has_decimal[0] and not col_has_decimal[1]:
                start_key = 6  # 第1列金额(下标6), 第2列数量(下标5)
            elif col_has_decimal[1] and not col_has_decimal[0]:
                start_key = 5  # 第1列数量(下标5), 第2列金额(下标6)
            else:
                start_key = 5  # 两列都整数或都小数 → 数量起(保守)
        elif ncols_data >= 2:
            start_key = 2  # 多列从名称开始(下标2=名称)
        elif all(_looks_like_number(c) for r in rows for c in r if c.strip()):
            # 单列纯数字
            has_decimal = any("." in c or "．" in c for r in rows for c in r if c.strip())
            if has_decimal:
                start_key = 6  # 金额(含小数)
            else:
                # 全整数: 检查最大位数, >4位(订单号/长编码) → 备注
                max_digits = 0
                for r in rows:
                    for c in r:
                        s = str(c).strip().replace(",", "").replace("，", "")
                        digits_only = re.sub(r"\D", "", s)  # 去掉所有非数字(含中间连字符)
                        if digits_only:
                            max_digits = max(max_digits, len(digits_only))
                if max_digits > 4:
                    start_key = 7  # 备注(长数字如订单号)
                else:
                    start_key = 5  # 数量(小整数)
        else:
            start_key = 2  # 单列文字→名称(下标2)

        # 需要粘贴的总行数(扩展)
        need = start_row + len(rows)
        while len(self.invoices) < need:
            self.invoices.append({"invoice_type": self.inv_type.get(), "tax_included": self.inv_taxinc.get(),
                                  "buyer": "", "tax_id": "", "is_natural": "",
                                  "qty": "", "amount": "", "remark": ""})
        self.invoice_refresh_tree()

        # 逐格写入
        # 特殊: 两列纯数字(金额+数量) → 直接映射列(不靠偏移)
        twocol_numeric = (ncols_data == 2 and start_key in (5, 6)
                          and all(all(_looks_like_number(c) for c in r if c.strip()) for r in rows))
        if twocol_numeric:
            # 第1列/第2列 → 数量(5) 或 金额(6), 按含小数分配
            col_has_dec = [False, False]
            for r in rows:
                for j in (0, 1):
                    if j < len(r) and r[j].strip():
                        if "." in r[j] or "．" in r[j]:
                            col_has_dec[j] = True
            for i, row in enumerate(rows):
                idx = start_row + i
                if idx >= len(self.invoices):
                    break
                for j in (0, 1):
                    if j >= len(row):
                        continue
                    val = row[j].strip()
                    if not val:
                        continue  # 跳过空值, 不覆盖已有
                    if col_has_dec[j]:
                        self.invoices[idx]["amount"] = val
                    else:
                        self.invoices[idx]["qty"] = val
        else:
            for i, row in enumerate(rows):
                for j, cell in enumerate(row):
                    k = start_key + j
                    if k >= len(key_cols):
                        break
                    idx = start_row + i
                    if idx >= len(self.invoices):
                        break
                    val = cell.strip()
                    key = key_cols[k]
                    if key == "is_natural":
                        self.invoices[idx][key] = "是" if val == "是" else ""
                    else:
                        self.invoices[idx][key] = val
        self.invoice_refresh_tree()
        # 粘贴完成: 清除点选状态(恢复到自动识别, 避免下次错乱)
        had_click = self._inv_click_col is not None
        self._inv_click_col = None
        self._inv_click_row = None
        self._inv_click_row_id = None
        tip = ", 下一点选可指定位置" if had_click else ""
        self.log(self.inv_log, f"✔ 已粘贴 {len(rows)} 行 × {ncols_data} 列{tip}")
        return "break"

    def invoice_refresh_tree(self):
        """数据模型 → 树(按流水号排序显示)"""
        self.inv_tree.delete(*self.inv_tree.get_children())
        for i, inv in enumerate(self.invoices, 1):
            self.inv_tree.insert("", "end", values=(
                f"{i:03d}",
                inv.get("invoice_type", "") or self.inv_type.get(),
                inv.get("tax_included", "") or self.inv_taxinc.get(),
                inv.get("buyer", ""), inv.get("tax_id", ""),
                inv.get("is_natural", "") or "", inv.get("qty", ""),
                inv.get("amount", ""), inv.get("remark", "")))
        # 斑马纹(网格横线感)
        if getattr(self, "_inv_zebra", None):
            self._inv_zebra()

    def invoice_add_row(self):
        """新增一行(空发票), 自动进入编辑"""
        self.invoices.append({"invoice_type": self.inv_type.get(), "tax_included": self.inv_taxinc.get(),
                              "buyer": "", "tax_id": "", "is_natural": "",
                              "qty": "", "amount": "", "remark": ""})
        self.invoice_refresh_tree()
        children = self.inv_tree.get_children()
        if children:
            last = children[-1]
            self.inv_tree.selection_set(last)
            self.inv_tree.see(last)
            bbox = self.inv_tree.bbox(last, "#2")  # 名称列
            if bbox:
                x, y = bbox[0] + 2, bbox[1] + 2
                self.inv_tree.event_generate("<Double-1>", x=x, y=y)
        self.log(self.inv_log, f"✔ 新增第 {len(self.invoices):03d} 行, 双击单元格填写或直接粘贴")

    def invoice_import_detail(self):
        """📥 导入发票明细文档(Excel/CSV), 自动映射列并加入发票列表
        映射: 发票抬头→购买方名称, 企业税号→纳税人识别号,
              发票金额→金额, 商品数量→数量
              抬头类型=个人→自然人"是"
        """
        path = filedialog.askopenfilename(
            title="选择发票明细文档",
            filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.csv"), ("所有文件", "*.*")])
        if not path:
            return
        try:
            rows = self._read_detail(path)
            if not rows:
                _mtop('showwarning', "提示", "文档无数据或无法识别表头")
                return
            head = rows[0]
            headers = [str(h or "").strip() for h in head]
            col_map = self._match_detail_columns(headers)
            missing = [n for n, ci in col_map.items() if ci is None]
            if missing:
                _mtop('showwarning', "提示", f"以下列未识别(请检查表头):\n{', '.join(missing)}")
            n = 0
            for r in rows[1:]:
                buyer = str(r[col_map["buyer"]] or "").strip() if col_map["buyer"] is not None else ""
                if not buyer:
                    continue
                inv = {
                    "invoice_type": self.inv_type.get(),
                    "tax_included": self.inv_taxinc.get(),
                    "buyer": buyer,
                    "tax_id": str(r[col_map["tax_id"]] or "").strip() if col_map["tax_id"] is not None else "",
                    "is_natural": "",
                    "qty": str(r[col_map["qty"]] or "").strip() if col_map["qty"] is not None else "",
                    "amount": str(r[col_map["amount"]] or "").strip() if col_map["amount"] is not None else "",
                    "remark": "",
                }
                if col_map["type"] is not None:
                    t = str(r[col_map["type"]] or "").strip()
                    if t and ("个人" in t or "自然人" in t):
                        inv["is_natural"] = "是"
                self.invoices.append(inv)
                n += 1
            self.invoice_refresh_tree()
            self.log(self.inv_log, f"✅ 已导入 {n} 条发票明细: {os.path.basename(path)}")
            _mtop('showinfo', "成功", f"已导入 {n} 条发票明细")
        except Exception as e:
            self.log(self.inv_log, f"❌ 导入失败: {e}")
            _mtop('showerror', "错误", f"导入失败:\n{e}")

    def _read_detail(self, path):
        """读取明细文档(csv/xlsx)为二维列表, 首行表头"""
        ext = os.path.splitext(path)[1].lower()
        if ext == ".csv":
            import csv as _csv
            with open(path, "r", encoding="utf-8-sig", errors="ignore") as f:
                return list(_csv.reader(f))
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return [[c.value for c in row] for row in ws.iter_rows()]

    def _match_detail_columns(self, headers):
        """按关键词匹配表头列(模糊)"""
        rules = {
            "buyer": ["抬头", "购买方名称", "购方名称", "买方名称", "发票抬头"],
            "tax_id": ["税号", "识别号", "纳税人"],
            "amount": ["发票金额", "金额", "价税合计", "合计金额"],
            "qty": ["商品数量", "数量"],
            "type": ["抬头类型", "购方类型", "买方类型", "类型"],
        }
        col_map = {}
        for key, kws in rules.items():
            found = None
            for i, h in enumerate(headers):
                hl = h.lower()
                if any(k.lower() in hl for k in kws):
                    found = i
                    break
            col_map[key] = found
        return col_map

    def invoice_auto_rows(self):
        """按粘贴的金额列自动生成行: 一键生成N行(每行一个金额)"""
        pass  # 粘贴已覆盖此功能

    def invoice_del_sel(self):
        sel = self.inv_tree.selection()
        if not sel:
            return
        idxs = sorted(self.inv_tree.index(i) for i in sel)
        for idx in reversed(idxs):
            if 0 <= idx < len(self.invoices):
                del self.invoices[idx]
        self.invoice_refresh_tree()

    def invoice_clear_col(self):
        """清空选中的某一列所有行的数据(使用上次点选的列或下拉指定列)"""
        # 优先: 点选的列; 其次: 下拉"单列粘贴"手动指定
        COL_KEY_MAP = {"发票类型": "invoice_type", "含税": "tax_included",
                       "名称": "buyer", "税号": "tax_id", "自然人": "is_natural",
                       "数量": "qty", "金额": "amount", "备注": "remark"}
        col_key = None
        col_label = None
        if self._inv_click_col is not None:
            # _inv_click_col: 0=发票类型 1=含税 2=名称 3=税号 4=自然人 5=数量 6=金额 7=备注
            key_names = ("invoice_type", "tax_included", "buyer", "tax_id", "is_natural",
                         "qty", "amount", "remark")
            labels = ("发票类型", "含税", "购买方名称", "纳税人识别号", "自然人", "数量", "金额", "备注")
            if 0 <= self._inv_click_col < len(key_names):
                col_key = key_names[self._inv_click_col]
                col_label = labels[self._inv_click_col]
        if col_key is None:
            sel_label = self.inv_paste_col.get().strip()
            col_key = COL_KEY_MAP.get(sel_label)
            col_label = sel_label
        if col_key is None or col_key == "serial":
            _mtop('showinfo', "提示", "请先点选要清空的列(单击该列任意一格), 或在\"单列粘贴→\"下拉选择")
            return
        if not self.invoices:
            _mtop('showwarning', "提示", "列表为空")
            return
        if not _mtop('askyesno', "确认", f"确定清空「{col_label}」列全部 {len(self.invoices)} 行数据?"):
            return
        for inv in self.invoices:
            inv[col_key] = ""
        self.invoice_refresh_tree()
        # 清空点选状态
        self._inv_click_col = None
        self._inv_click_row = None
        self._inv_click_row_id = None
        self.log(self.inv_log, f"🗑 已清空「{col_label}」列全部 {len(self.invoices)} 行")

    def invoice_clear(self):
        if self.invoices and _mtop('askyesno', "确认", "清空全部发票?"):
            self.invoices.clear()
            self.invoice_refresh_tree()
            # 清空时重置点选粘贴状态, 避免下次粘贴残留到旧列
            self._inv_click_col = None
            self._inv_click_row = None
            self._inv_click_row_id = None
            self.log(self.inv_log, "🗑 列表已清空 (粘贴已恢复自动识别)")

    def invoice_bulk_dlg(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("批量粘贴导入")
        dlg.geometry("600x430")
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.attributes("-topmost", True)  # 系统顶层
        ttk.Label(dlg, text="每行一条发票, 格式: 购买方名称, 税号或\"是\", 数量, 金额 (,备注可选)\n"
                            "例: 张三,是,2,52.20    或  公司A,91370306MA3TBJ0T8E,1,122.00,加急\n"
                            "回车粘贴后点\"导入\"").pack(anchor="w", padx=10, pady=5)
        txt = scrolledtext.ScrolledText(dlg, height=14, font=("Consolas", 10))
        txt.pack(fill="both", expand=True, padx=10, pady=5)

        def do_import():
            try:
                lst = parse_bulk_text(txt.get("1.0", "end"))
            except ValueError as e:
                _mtop('showerror', "格式错误", str(e), parent=dlg)
                return
            if not lst:
                _mtop('showwarning', "提示", "没有解析到任何行", parent=dlg)
                return
            self.invoices.extend(lst)
            self.invoice_refresh_tree()
            self.log(self.inv_log, f"✔ 批量导入 {len(lst)} 条")
            dlg.destroy()

        frm = ttk.Frame(dlg)
        frm.pack(fill="x", padx=10, pady=8)
        ttk.Button(frm, text="导入", command=do_import).pack(side="right", padx=5)
        ttk.Button(frm, text="取消", command=dlg.destroy).pack(side="right", padx=5)

    def invoice_generate(self):
        if not self.invoices:
            _mtop('showwarning', "提示", "发票列表为空, 请先添加")
            return
        self._inv_close_editor()
        # 用当前固定内容覆盖默认
        fixed = {
            "invoice_type": self.inv_type.get(),
            "tax_included": self.inv_taxinc.get(),
            "item_name": self.inv_item.get().strip(),
            "tax_code": self.inv_code.get().strip(),
            "unit": self.inv_unit.get().strip(),
            "tax_rate": self.inv_rate.get().strip(),
        }
        # 行级值优先, 空则用固定内容默认
        invoices = []
        for inv in self.invoices:
            merged = dict(inv)
            for k, v in fixed.items():
                if not str(merged.get(k, "") or "").strip():
                    merged[k] = v
            invoices.append(merged)
        tpl = self.inv_tpl.get().strip().strip('"') or find_template()

        out = filedialog.asksaveasfilename(
            defaultextension=".xlsx", initialfile="开票导入.xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if not out:
            return
        self.inv_log.insert("end", "▶ 生成中...\n")
        try:
            path, errs = generate_invoice_xlsx(invoices, template_path=tpl, out_path=out)
            if errs:
                self.log(self.inv_log, f"❌ 校验未通过 ({len(errs)} 条):")
                for e in errs:
                    self.log(self.inv_log, f"   {e}")
                _mtop('showerror', "校验失败", "请修正数据:\n" + "\n".join(errs[:10]))
                return
            self.log(self.inv_log, f"✅ 生成成功: {path}")
            self.log(self.inv_log, f"   共 {len(invoices)} 张发票, 流水号 001~{len(invoices):03d}")
            _mtop('showinfo', "成功", f"开票文件已生成!\n{path}\n共 {len(invoices)} 张")
        except Exception as e:
            self.log(self.inv_log, f"❌ 错误: {e}")
            _mtop('showerror', "错误", str(e))

    # ---------- 内置更新 ----------
    def check_update_btn(self):
        """手动检查更新按钮 (后台线程, 不阻塞UI; 防重入)"""
        if not UPDATE_AVAILABLE:
            _mtop('showinfo', "提示", "更新模块未加载")
            return
        if getattr(self, "_checking", False):
            self.set_global_msg("⏳ 正在检查更新中, 请稍候...")
            return
        self._checking = True
        self.set_global_msg("🔍 检查更新中...")
        self.update_btn_hint()
        import threading

        def _work():
            try:
                asset = self_asset_name()
                info = check_update(asset_name=asset)
                self.root.after(0, lambda: self._check_update_done(info))
            except Exception as e:
                self.root.after(0, lambda: self._check_update_done(None, str(e)))

        t = threading.Thread(target=_work, daemon=True)
        t.start()

    def set_global_msg(self, text):
        """全局状态信息(所有Tab可见), 不再写到开票页log"""
        try:
            self.status_msg.set(text)
        except Exception:
            pass

    def update_btn_hint(self):
        """检查更新按钮提示文字"""
        try:
            self.update_btn.config(text="🔍 检查更新 (v)" if not getattr(self, "_update_avail", False)
                                   else "✨ 发现新版本!")
        except Exception:
            pass

    def _check_update_done(self, info, err=None):
        self._checking = False
        if err:
            self.set_global_msg(f"❌ 更新检查错误: {err}")
            _mtop('showerror', "错误", str(err))
            return
        if info is None:
            self.set_global_msg("⚠️ 检查失败(网络/API不可达)")
            _mtop('showwarning', "提示", "检查更新失败, 请检查网络")
            return
        if not info["has_update"]:
            self.set_global_msg(f"✅ 已是最新版本 ({info['current_tag']})")
            _mtop('showinfo', "提示", f"已是最新版本 {info['current_tag']}")
            return
        # 有更新: 标记按钮 + 弹窗(仅手动检查时弹, 静默只标记)
        self._update_avail = True
        self.update_btn_hint()
        body = (info["body"] or "").strip()
        body_preview = "\n".join(body.splitlines()[:6]) if body else "(无更新日志)"
        if not _mtop('askyesno', "发现新版本",
                                   f"发现新版本 {info['latest_tag']}\n"
                                   f"当前版本 {info['current_tag']}\n\n"
                                   f"更新内容:\n{body_preview}\n\n"
                                   f"是否下载更新?"):
            return
        self.do_download_update(info)

    def do_download_update(self, info):
        """下载新exe → 启动updater替换 (目标固定标准名, 消除new_残留)"""
        url = info.get("download_url")
        if not url:
            _mtop('showerror', "错误", "未找到下载地址")
            return
        exe_name = self_exe_name()
        temp_dir = os.path.dirname(os.path.abspath(sys.executable if getattr(sys, "frozen", False) else __file__))
        # 下载临时名 = new_{标准资产名}, 替换目标 = 标准资产名(不管当前运行名)
        std_name = self_asset_name()
        new_exe = os.path.join(temp_dir, f"new_{std_name}")
        target = os.path.join(temp_dir, std_name)
        # 显示进度条
        try:
            self.dl_progress.pack(side="right", padx=8)
            self.dl_progress["value"] = 0
            self.dl_pct.set("0%")
        except Exception:
            pass

        def _p(got, total):
            # 线程中回调 → 主线程更新UI
            if total:
                pct = min(100, int(got * 100 / total))
                self.root.after(0, lambda: (self.dl_progress.configure(value=pct),
                                            self.dl_pct.set(f"{pct}%")))

        self.log(self.inv_log, f"⬇ 下载中: {info['latest_tag']} ({std_name})")
        # 后台线程下载, 避免阻塞UI
        import threading

        def _work():
            ok, size, err = download_file(url, new_exe, timeout=300, progress_cb=_p)
            self.root.after(0, lambda: self._download_done(info, ok, size, err, new_exe, exe_name, target, temp_dir))

        t = threading.Thread(target=_work, daemon=True)
        t.start()

    def _download_done(self, info, ok, size, err, new_exe, exe_name, target, temp_dir):
        # 隐藏进度条
        try:
            self.dl_progress.pack_forget()
        except Exception:
            pass
        if not ok:
            url = info.get("download_url", "")
            self.log(self.inv_log, f"❌ 下载失败: {err}")
            _mtop('showerror', "下载失败", f"请手动下载:\n{url}\n\n{err}")
            return
        self.log(self.inv_log, f"✅ 下载完成 ({size/1024/1024:.1f}MB), 准备更新...")
        # 生成 updater.bat 并启动 (替换目标=标准名)
        old_exe = exe_name
        bat = make_updater_bat(new_exe, old_exe, temp_dir, target=target)
        bat_path = os.path.join(temp_dir, "updater.bat")
        with open(bat_path, "w", encoding="ascii") as f:
            f.write(bat)
        self.log(self.inv_log, "🔄 准备更新...")
        _mtop('showinfo', "更新", "下载完成, 点确定后程序将自动更新并重启(约5秒)")

        def _do_update():
            # 确保GUI销毁+进程强制退出, updater才能接管替换
            try:
                run_updater(bat_path, new_exe, exe_name, temp_dir)
            except Exception:
                pass
            try:
                self.root.destroy()
            except Exception:
                pass
            try:
                import sys
                sys.stdout.flush()
            except Exception:
                pass
            os._exit(0)  # 强制退出, 不留存活进程

        self.root.after(1000, _do_update)

    def check_update_silent(self):
        """启动时后台静默检查(不打扰): 有更新只标记按钮, 不弹窗(防双弹窗)"""
        if not UPDATE_AVAILABLE:
            return

        def _do():
            try:
                asset = self_asset_name()
                info = check_update(asset_name=asset)
                if info and info["has_update"]:
                    # 只标记按钮提示, 不自动弹窗
                    self.root.after(0, lambda: self._mark_update_avail(info))
            except Exception:
                pass

        import threading
        t = threading.Thread(target=_do, daemon=True)
        t.start()

    def _mark_update_avail(self, info):
        self._update_avail = True
        if not getattr(self, "_checking", False):
            self.update_btn_hint()
            self.set_global_msg(f"✨ 发现新版本 {info['latest_tag']} — 点击右上角检查更新")

    def _show_silent_update(self, info):
        body = (info["body"] or "").strip()
        body_preview = "\n".join(body.splitlines()[:4]) if body else ""
        if _mtop('askyesno', "发现新版本(via 大萝北拔萝卜)",
                               f"发现新版本 {info['latest_tag']} (当前 {info['current_tag']})\n\n"
                               f"{body_preview}\n\n是否更新?"):
            self.do_download_update(info)

    # ---------- 通用 ----------
    def pick(self, var):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm"), ("所有", "*.*")])
        if p:
            var.set(p)

    def pick_dir(self, var):
        p = filedialog.askdirectory()
        if p:
            var.set(p)

    def _add_row(self, parent):
        row = tk.Frame(parent)
        idx = len(self.map_rows)
        s = tk.StringVar(value="规格商家编码" if idx == 0 else "")
        t = tk.StringVar(value="规格商家编码" if idx == 0 else "")
        ttk.Label(row, text=f"映射{idx + 1}:").pack(side="left")
        ttk.Entry(row, textvariable=s, width=16).pack(side="left", padx=3)
        ttk.Label(row, text="→").pack(side="left")
        ttk.Entry(row, textvariable=t, width=16).pack(side="left", padx=3)
        row.pack(fill="x", padx=5, pady=2)
        self.map_rows.append((s, t))

    def log(self, box, txt):
        box.insert("end", txt + "\n")
        box.see("end")

    def run(self):
        src = self.src_var.get().strip().strip('"')
        tgt = self.tgt_var.get().strip().strip('"')
        if not src or not tgt:
            _mtop('showwarning', "提示", "请先选择源表和目标表")
            return
        fill_map = [(s.get().strip(), t.get().strip())
                    for s, t in self.map_rows if s.get().strip() and t.get().strip()]
        if not fill_map:
            _mtop('showwarning', "提示", "请至少填写一个回填映射")
            return
        self.run_btn.config(state="disabled")
        self.log(self.match_log, "▶ 开始核对...")
        try:
            matched, multi, notfound, out_path = run_match(
                src, tgt, self.sk.get().strip(), self.tk.get().strip(),
                fill_map, self.sku_col.get().strip() or None,
                self.rmk_col.get().strip() or None, self.skip_existing.get())
            self.log(self.match_log, f"✅ 匹配: {matched}, 多规格: {multi}, 未匹配: {len(notfound)}")
            self.log(self.match_log, f"   已另存(源文件未改): {os.path.basename(out_path)}")
            if notfound:
                self.log(self.match_log, f"\n⚠️ 未匹配 ({len(notfound)}):")
                for v in notfound[:50]:
                    self.log(self.match_log, f"   {v}")
                if len(notfound) > 50:
                    self.log(self.match_log, f"   ... 等 {len(notfound)} 个")
            _mtop('showinfo', "成功", f"核对完成!\n匹配 {matched}\n多规格 {multi}\n未匹配 {len(notfound)}")
        except Exception as e:
            msg = str(e)
            self.log(self.match_log, f"❌ 错误: {msg}")
            if "Fill" in msg:
                _mtop('showerror', "文件格式问题",
                      "表格文件样式异常(不兼容的填充格式)。\n\n"
                      "解决办法: 用 WPS/Excel 打开该文件 → 另存为 .xlsx 后重试。\n\n"
                      f"详细: {msg}")
            else:
                _mtop('showerror', "错误", msg)
        finally:
            self.run_btn.config(state="normal")

    def run_img(self):
        src = self.img_src.get().strip().strip('"')
        if not src:
            _mtop('showwarning', "提示", "请选择源文件夹")
            return
        q = max(1, min(100, self.img_q.get()))
        global QUALITY
        QUALITY = q
        if self.img_out.get().strip():
            out = self.img_out.get().strip().strip('"')
        else:
            out = src.rstrip('/\\') + "_jpg"
        self.img_btn.config(state="disabled")
        self.log(self.img_log, f"▶ 转换中... 源: {src} → {out}")
        try:
            total, converted, copied, failed, errors = run_imgconvert(src, out)
            self.log(self.img_log, f"✅ 完成: 总{total}, 转JPG{converted}, 复制{copied}, 失败{failed}")
            self.log(self.img_log, f"   输出: {out}")
            if errors:
                self.log(self.img_log, "⚠️ 失败清单:")
                for rel, e in errors[:10]:
                    self.log(self.img_log, f"   {rel}: {e}")
            _mtop('showinfo', "成功", f"转换完成!\n总{total}\n转JPG {converted}\n复制 {copied}\n失败 {failed}")
        except Exception as e:
            self.log(self.img_log, f"❌ 错误: {e}")
            _mtop('showerror', "错误", str(e))
        finally:
            self.img_btn.config(state="normal")


def main():
    root = tk.Tk()
    # 设置窗口/任务栏图标 (打包时 --add-data 附带 app.ico)
    try:
        if getattr(sys, "frozen", False):
            icon_path = os.path.join(sys._MEIPASS, "app.ico")
        else:
            icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "app.ico")
        if os.path.exists(icon_path):
            root.iconbitmap(icon_path)
    except Exception:
        pass
    try:
        ttk.Style().theme_use("clam")
    except Exception:
        pass
    app = App(root)
    _set_mbox_root(root)
    # 启动后后台静默检查更新(有新版才提示)
    try:
        app.check_update_silent()
    except Exception:
        pass
    root.mainloop()


if __name__ == "__main__":
    main()
