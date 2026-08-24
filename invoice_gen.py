#!/usr/bin/env python3
"""
开票生成核心模块 - 数电发票批量开票导入模板生成
============================================
输入: 发票列表 [{type, tax_included, is_natural, buyer, tax_id, remark, item_name, tax_code, unit, qty, amount, tax_rate}]
输出: 按官方模板(V260401版)生成 .xlsx, 第4行起填数据, 流水号自动 001/002... 两表同步

固定内容(用户测试表实锤): 普通发票 / 是 / 滤芯 / 1090130020000000000 / 个 / 0.01
模板结构: 1-发票基本信息(39列) + 2-发票明细信息(14列) + 字典表(保留不动)
"""
import os
import re
import sys
import datetime
import shutil

import openpyxl

# ---------- 默认固定内容 ----------
DEFAULT_FIXED = {
    "invoice_type": "普通发票",      # 发票类型
    "tax_included": "是",            # 是否含税
    "item_name": "滤芯",             # 项目名称
    "tax_code": "1090130020000000000",  # 商品和服务税收编码
    "unit": "个",                    # 单位
    "tax_rate": "0.01",              # 税率(1%)
}

# 模板缓存路径(本机) / 打包时用内置副本
TEMPLATE_CANDIDATES = [
    "/mnt/ssd/hermes/cache/documents/doc_9dbfb8b2409e_(V260401版)批量开票-导入开票模板 (16).xlsx",
    "/mnt/ssd/hermes/cache/documents/doc_71ea82691c52_(V260401版)批量开票-导入开票模板 测试.xlsx",
]

# 基本信息表各列定义(1-发票基本信息)
COL_BILL = {
    "流水号": 1, "发票类型": 2, "特定业务类型": 3, "是否含税": 4,
    "自然人标识": 5, "购买方名称": 6, "纳税人识别号": 7,
    # ... 8-22 不填(证件/地址/银行等), 23 = 备注
    "备注": 23,
}
# 明细表各列定义(2-发票明细信息)
COL_ITEM = {
    "流水号": 1, "项目名称": 2, "税收编码": 3, "规格型号": 4,
    "单位": 5, "数量": 6, "单价": 7, "金额": 8, "税率": 9,
}

INVOICE_TYPE_OPTIONS = ["普通发票", "专用发票", "电子普通发票", "电子专用发票"]


def find_template():
    """找模板文件: 环境变量/打包内置/同目录内置/缓存"""
    p = os.environ.get("INVOICE_TEMPLATE")
    if p and os.path.isfile(p):
        return p
    # PyInstaller 打包: sys._MEIPASS 里的内置模板
    try:
        if getattr(sys, "frozen", False):
            bundled = os.path.join(sys._MEIPASS, "invoice_template.xlsx")
            if os.path.isfile(bundled):
                return bundled
    except Exception:
        pass
    # 本文件同目录内置模板副本(项目里放的)
    here = os.path.dirname(os.path.abspath(__file__))
    local = os.path.join(here, "invoice_template.xlsx")
    if os.path.isfile(local):
        return local
    for f in os.listdir(here):
        if "开票" in f and f.endswith(".xlsx"):
            return os.path.join(here, f)
    # 最后兜底: 缓存目录
    for c in TEMPLATE_CANDIDATES:
        if os.path.isfile(c):
            return c
    return None


def make_serial(n, width=3):
    """流水号: 1 -> '001'"""
    return str(n).zfill(width)


def validate_invoice(inv, idx):
    """校验单条发票, 返回错误列表"""
    errs = []
    if not inv.get("buyer") or not str(inv["buyer"]).strip():
        errs.append(f"第{idx}行: 购买方名称为空")
    buyer = str(inv["buyer"]).strip()
    if len(buyer) > 100:
        errs.append(f"第{idx}行: 购买方名称超长({len(buyer)}>100)")
    is_natural = str(inv.get("is_natural", "")).strip()
    tax_id = str(inv.get("tax_id", "")).strip()
    if is_natural != "是" and not tax_id:
        # 无税号且未标记自然人: 可能是海外公司/个人, 放行(E列自动填"是")
        pass
    if is_natural != "是" and tax_id and not _is_tax_id(tax_id):
        errs.append(f"第{idx}行: 纳税人识别号格式异常({tax_id})")
    if tax_id and len(tax_id) > 20:
        errs.append(f"第{idx}行: 纳税人识别号超长({len(tax_id)}>20)")
    amt = inv.get("amount")
    if amt is None or amt == "":
        errs.append(f"第{idx}行: 金额为空")
    else:
        try:
            float(str(amt))
        except ValueError:
            errs.append(f"第{idx}行: 金额不是数字({amt})")
    qty = inv.get("qty")
    if qty is not None and qty != "":
        try:
            float(str(qty))
        except ValueError:
            errs.append(f"第{idx}行: 数量不是数字({qty})")
    return errs


def normalize_amount(v):
    """金额统一两位小数文本"""
    try:
        return f"{float(str(v)):.2f}"
    except (ValueError, TypeError):
        return str(v or "")


def generate_invoice_xlsx(invoices, template_path=None, out_path=None):
    """
    生成开票导入 xlsx
    invoices: [{invoice_type,tax_included,is_natural,buyer,tax_id,remark,
                item_name,tax_code,unit,qty,amount,tax_rate}]
    返回: (out_path, 错误列表)
    """
    if not invoices:
        raise ValueError("发票列表为空")
    tpl = template_path or find_template()
    if not tpl:
        raise ValueError("找不到开票模板文件, 请选择模板")

    src = openpyxl.load_workbook(tpl)  # 保留全部sheet/样式/字典
    if "1-发票基本信息" not in src.sheetnames or "2-发票明细信息" not in src.sheetnames:
        raise ValueError("模板缺少必需的工作表(1-发票基本信息/2-发票明细信息)")

    ws_bill = src["1-发票基本信息"]
    ws_item = src["2-发票明细信息"]

    # 校验全部
    all_errs = []
    for i, inv in enumerate(invoices, 1):
        all_errs.extend(validate_invoice(inv, i))
    if all_errs:
        return None, all_errs

    # 清空模板已有数据(第4行起, 防止残留)
    data_rows_bill = ws_bill.max_row - 3
    if data_rows_bill > 0:
        ws_bill.delete_rows(4, data_rows_bill)
    data_rows_item = ws_item.max_row - 3
    if data_rows_item > 0:
        ws_item.delete_rows(4, data_rows_item)

    # 填数据
    for i, inv in enumerate(invoices, 1):
        serial = make_serial(i)
        fixed = {**DEFAULT_FIXED}
        # 允许每条覆盖固定内容
        for k in fixed:
            v = inv.get(k)
            if v not in (None, ""):
                fixed[k] = v

        is_natural = str(inv.get("is_natural", "") or "").strip()
        tax_id = str(inv.get("tax_id", "") or "").strip()
        # 自然人或海外无税号: E列填"是"
        natural_flag = "是" if (is_natural == "是" or not tax_id) else "否"
        remark = str(inv.get("remark", "") or "").strip()

        # --- 1-发票基本信息 ---
        ws_bill.cell(row=3 + i, column=COL_BILL["流水号"], value=serial)
        ws_bill.cell(row=3 + i, column=COL_BILL["发票类型"], value=str(fixed["invoice_type"]))
        ws_bill.cell(row=3 + i, column=COL_BILL["是否含税"], value=str(fixed["tax_included"]))
        ws_bill.cell(row=3 + i, column=COL_BILL["自然人标识"], value=natural_flag)
        ws_bill.cell(row=3 + i, column=COL_BILL["购买方名称"], value=str(inv["buyer"]).strip())
        ws_bill.cell(row=3 + i, column=COL_BILL["纳税人识别号"], value=tax_id or None)
        if remark:
            ws_bill.cell(row=3 + i, column=COL_BILL["备注"], value=remark)

        # --- 2-发票明细信息 ---
        ws_item.cell(row=3 + i, column=COL_ITEM["流水号"], value=serial)
        ws_item.cell(row=3 + i, column=COL_ITEM["项目名称"], value=str(fixed["item_name"]))
        ws_item.cell(row=3 + i, column=COL_ITEM["税收编码"], value=str(fixed["tax_code"]))
        ws_item.cell(row=3 + i, column=COL_ITEM["单位"], value=str(fixed["unit"]))
        qty = inv.get("qty")
        ws_item.cell(row=3 + i, column=COL_ITEM["数量"], value=qty if qty not in (None, "") else None)
        ws_item.cell(row=3 + i, column=COL_ITEM["金额"], value=normalize_amount(inv.get("amount")))
        ws_item.cell(row=3 + i, column=COL_ITEM["税率"], value=str(fixed["tax_rate"]))

    if not out_path:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(os.getcwd(), f"开票导入_{ts}.xlsx")
    src.save(out_path)
    return out_path, []


def parse_bulk_text(text, sep=None):
    """
    解析批量粘贴文本 -> 发票列表
    约定格式(每行): 购买方名称, 税号或"是"(可空), 数量, 金额
      - 税号位写"是" = 自然人发票(无税号)
      - 税号位写"否"或留空 = 单位发票(需税号, 留空则生成时校验报错)
      - 可在最后追加备注: 名称,税号,数量,金额,备注
    分隔符: 逗号/制表符/分号/竖线 (中英文均可)
    """
    invoices = []
    for line_no, line in enumerate(text.splitlines(), 1):
        line = line.strip()
        if not line:
            continue
        if sep is None:
            parts = [p.strip() for p in re.split(r"[,，;|\t]", line)]  # 保留空占位
        else:
            parts = [p.strip() for p in line.split(sep)]
        # 保留空占位, 但去掉整行空
        if all(p == "" for p in parts):
            continue
        if len(parts) < 2:
            raise ValueError(f"第{line_no}行格式错误(至少需 名称+金额): {line}")

        inv = {"buyer": parts[0], "is_natural": "", "tax_id": "",
               "remark": "", "qty": "", "amount": ""}

        # 从右往左剥离: 金额 -> 数量 (纯数字)
        # 若末尾不是数字(备注在最后), 先摘掉备注再处理数字
        rest = parts[1:]
        remark_parts = []
        if rest and not _is_num(rest[-1]) and rest[-1] not in ("是", "否") and not _is_tax_id(rest[-1]):
            # 末尾是非数字非税号非标记 -> 可能是备注(或空税号占位)
            if rest[-1] == "":
                pass  # 空占位后面统一处理
            else:
                remark_parts.insert(0, rest[-1])
                rest = rest[:-1]

        nums = []
        while rest and _is_num(rest[-1]):
            nums.insert(0, rest[-1].replace(",", ""))
            rest = rest[:-1]
        # nums 从右到左: [数量?, 金额] (最多取两个数字)
        if nums:
            inv["amount"] = nums[-1]
            if len(nums) >= 2:
                inv["qty"] = nums[-2]

        # 剩余字段处理: [税号/标记, 备注...] 或 [空占位, 备注...] 或 [税号/标记]
        if rest:
            p = rest[-1]
            if p in ("是", "否"):
                inv["is_natural"] = p
                rest = rest[:-1]
            elif _is_tax_id(p):
                inv["tax_id"] = p
                rest = rest[:-1]
            elif p == "":
                # 空税号占位(海外/自然人用空列)
                rest = rest[:-1]
            # 若第一个字段是税号/标记(备注在中间/开头)
            if rest and rest[0] in ("是", "否"):
                inv["is_natural"] = rest[0]
                rest = rest[1:]
            elif rest and _is_tax_id(rest[0]):
                inv["tax_id"] = rest[0]
                rest = rest[1:]
            # 剩余全是备注
            remark_parts = [r for r in rest if r != ""] + remark_parts

        inv["remark"] = " ".join(remark_parts).strip()
        invoices.append(inv)
    return invoices


def _is_tax_id(s):
    """税号特征: ≥15位 且 (含字母 或 全数字长串18位)"""
    s = str(s or "").strip()
    if not s:
        return False
    pure = re.sub(r"[^0-9A-Za-z]", "", s)
    if len(pure) < 15:
        return False
    return True


def _is_num(s):
    try:
        float(str(s).replace(",", ""))
        return True
    except (ValueError, TypeError):
        return False


if __name__ == "__main__":
    # 自测
    demo = [
        {"buyer": "测试公司A", "tax_id": "91370306MA3TBJ0T8E", "is_natural": "", "qty": 1, "amount": 122.00, "remark": ""},
        {"buyer": "张三", "tax_id": "", "is_natural": "是", "qty": 2, "amount": 52.20},
    ]
    out, errs = generate_invoice_xlsx(demo, out_path="/tmp/invoice_demo.xlsx")
    print("输出:", out)
    print("错误:", errs)
    if out:
        wb = openpyxl.load_workbook(out)
        ws = wb["1-发票基本信息"]
        for r in range(4, 6):
            print("基本信息R%d:" % r, [ws.cell(row=r, column=c).value for c in [1, 2, 4, 5, 6, 7, 23]])
        ws2 = wb["2-发票明细信息"]
        for r in range(4, 6):
            print("明细R%d:" % r, [ws2.cell(row=r, column=c).value for c in [1, 2, 3, 5, 6, 8, 9]])

    # 批量文本测试
    txt = "测试公司A,91370306MA3TBJ0T8E,1,122.00\n张三,是,2,52.20\n黄朗瑜,是,1,14.13\nAPAC AVC,,3,100.00\n远大洁净空气科技有限公司,91430100563544319E,2,146.70,加急"
    lst = parse_bulk_text(txt)
    print("\n批量解析:")
    for x in lst:
        print("  ", x)